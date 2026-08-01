import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEXT_COLUMN = 'comment_body_clean_final'
HARD_EXACT_PATTERNS = ['^/?html_details_block\\.?$', '^/?html_image_reference\\.?$', '^/?html_meta_comment\\.?$', '^/?code_block_attached\\.?$', '^/?diff_attached\\.?$', '^/?error_log_attached\\.?$', '^/?log_reference\\.?$', '^/?hash_reference\\.?$', '^/?url_reference\\.?$', '^/?benchmark_reference\\.?$', '^/?ci_status_reference\\.?$', '^/?workflow_ci\\.?$', '^/?workflow_review\\.?$', '^/?workflow_coordination\\.?$', '^/?approval_notification\\.?$', '^/?automated_cla_check\\.?$', '^/?automated_bot_message\\.?$', '^/?inline_code\\.?$', '^a github user is mentioned\\.?$', '^multiple github users are mentioned\\.?$', '^a github user is asked to review or check this\\.?$', '^multiple github users are asked to review or check this\\.?$', '^/[a-zA-Z0-9_-]+$', '^/[a-zA-Z0-9_-]+\\s+[a-zA-Z0-9_.-]+$', '^[a-f0-9]{7,40}$', '^\\s*$', '^\\?+$', '^!+$', '^\\)+$', '^\\(+$', '^eyes:?$', '^also$', '^node$', '^\\d+$']
HARD_PHRASE_PATTERNS = ['^landed in\\s+/?hash_reference\\.?$', '^fixed in\\s+/?hash_reference\\.?$', '^merged in\\s+/?hash_reference\\.?$', '^closed by\\s+/?hash_reference\\.?$', 'the following test[s]? \\*\\*?failed\\*\\*?', 'full pr test history', 'your pr dashboard', 'rerun all failed tests', 'rerun all mandatory failed tests', 'please help us cut down on flakes', 'unknown cla label state', 'codecov .* report', '^lgtm label has been added\\.?\\s*/?html_details_block\\.?$', '^approved label has been added\\.?\\s*/?html_details_block\\.?$', '^pull-request has been approved\\.?\\s*/?html_details_block\\.?$', '^pull request has been approved\\.?\\s*/?html_details_block\\.?$', '^[a-z0-9_-]+ label has been added\\.?\\s*/?html_details_block\\.?$', '^[a-z0-9_-]+ label has been removed\\.?\\s*/?html_details_block\\.?$', 'new changes are detected\\.?\\s*lgtm label has been (added|removed)', 'new changes are detected\\.?\\s*lgtm label has been', 'label\\(s\\).*cannot be applied', 'the label .* cannot be applied', "repository doesn't have them", 'only github organization members can add the label', 'you can only set the release note label', 'release-note-label-needed', 'no release-note block was detected', "github didn't allow me to request pr reviews", 'failed to re-open pr', 'state cannot be changed', 'those labels are not set on the issue', 'there are no sig labels on this issue', 'adding label .* merge commits', 'release-note-edit must be used with a release note block', 'this request has been marked as needing help from a contributor', 'this request has been marked as suitable for new contributors', 'guidelines\\s+please ensure that the issue body includes', 'instructions for interacting with me using pr comments', 'verify that this patch is reasonable to test', 'i will not automatically test new commits in this pr', '^a github user is mentioned\\.?\\s*!cat image\\s*/?url_reference\\s*/?html_details_block\\.?$', '^multiple github users are mentioned\\.?\\s*!cat image\\s*/?url_reference\\s*/?html_details_block\\.?$', '!cat image\\s*/?url_reference', 'this issue is currently awaiting triage', 'if a sig or subproject determines this is a relevant issue']
PLACEHOLDERS_TO_REMOVE = ['/html_details_block', '/html_image_reference', '/html_meta_comment', '/code_block_attached', '/diff_attached', '/error_log_attached', '/log_reference', '/hash_reference', '/url_reference', '/benchmark_reference', '/ci_status_reference', '/automated_cla_check', '/automated_bot_message', '/approval_notification', '/workflow_ci', '/workflow_review', '/workflow_coordination', '/inline_code', '/emoji_celebra', '/emoji_other']
GENERIC_USER_PLACEHOLDERS = ['a github user is mentioned\\.?', 'multiple github users are mentioned\\.?', 'a github user is asked to review or check this\\.?', 'multiple github users are asked to review or check this\\.?']
LOW_VALUE_VISUAL_PATTERNS = ['!cat image', '!dog image', '!gif', '!meme']
AUTOMATED_PREFIX_MARKERS = ['/automated_bot_message', 'automated_bot_message', '/approval_notification', 'approval_notification', '/automated_cla_check', 'automated_cla_check', '/workflow_ci', 'workflow_ci', '/workflow_review', 'workflow_review', '/workflow_coordination', 'workflow_coordination', '/ci_status_reference', 'ci_status_reference']
OPERATIONAL_EXACT_PATTERNS = ['^thanks\\.?$', '^thanks!$', '^thank you\\.?$', '^ok\\.?$', '^okay\\.?$', '^yes\\.?$', '^yeah\\.?$', '^no worries\\.?$', '^sounds good\\.?$', '^looks good\\.?$', '^all good!?$', '^acknowledged\\.?$', '^done\\.?$', '^updated\\.?$', '^fixed\\.?$', '^resolved\\.?$', '^addressed\\.?$', '^squashed\\.?$', '^rebased\\.?$', '^rebase\\.?$', '^rebase done\\.?$', '^force-pushed\\.?$', '^amended\\.?$', '^ptal\\.?$', '^lgtm\\.?$', '^approved\\.?$', '^approve, thanks!?$', '^looks good to me\\.?$', '^this one is looking good\\.?$', '^ready for review\\.?$', '^ready for another pass\\.?$', '^i am reviewing this pr\\.?$', '^reviewing new commits\\.?$', '^ci is green.*$', '^ci looks good.*$', '^test was successful!?$', '^tests? passed\\.?$', '^passes locally\\.?$', '^passed locally\\.?$', '^unrelated failure\\.?$', '^seems unrelated\\.?$', '^not a flaking test\\.?$']
OPERATIONAL_PHRASE_PATTERNS = ['^needs rebase\\.?$', '^please rebase\\.?$', '^rebased to fix conflict.*$', '^rebased and removed merge commit.*$', '^i have rebased.*$', '^resolved trivial conflict.*$', '^resolved conflict.*$', '^only solved conflicts.*$', '^merge conflict.*$', '^done,? thanks.*$', '^fixed now\\.?$', '^should be fixed now\\.?$', '^i have revised it\\.?$', '^i revised it\\.?$', '^i updated it\\.?$', '^i fixed it\\.?$', '^i pushed it\\.?$', '^squashed.*$', '^amended the commit.*$', '^force-pushed.*$', '^comments addressed\\.?$', '^all comments have been addressed\\.?$', '^adjusted all comments.*$', '^i think i addressed all the comments\\.?$', '^i have addressed all the comments\\.?$', '^i believe i have resolved all review comments\\.?$', '^i believe your comments have been addressed\\.?$', '^friendly ping.*$', '^gentle ping.*$', '^ping for review.*$', '^pinging .* for review.*$', '^please take a look\\.?$', '^can you take a look\\.?$', '^can you please take a look\\.?$', '^could you please take a look\\.?$', '^could you help review\\.?$', '^please review\\.?$', '^review,? thanks!?$', '^ptal\\.?$', '^ptal,? thanks!?$', '^any update\\??$', '^any further feedback\\??$', '^does anything else still need to be done\\??$', '^is this still active\\??$', '^can i work on this issue\\??$', '^can i fix\\??$', '^i want to work on this issue\\.?$', '^i would like to work on this\\.?$', '^i would like to work on this issue\\.?$', "^i'd like to work on this\\.?$", "^i'd like to work on this issue\\.?$", '^may i take this\\??$', '^may i take this issue\\??$', '^i will take this\\.?$', '^i will take this issue\\.?$', '^i will take up this issue\\.?$', "^i'll take this\\.?$", "^i'll take this issue\\.?$", '^i picked this up\\.?$', '^i picked up this issue\\.?$', "^i'm taking this\\.?$", "^i'm taking this issue\\.?$", '^i will take a look\\.?$', "^i'll take a look\\.?$", "^i'll give it a try\\.?$", "^i'll try\\.?$", '^i will try\\.?$', '^(ok|okay|sure|alright|fine),?\\s+i (will|can) do it( tomorrow| later)?\\.?$', '^i (will|can) do it( tomorrow| later)?\\.?$', "^i'll do it( tomorrow| later)?\\.?$", '^based on the quotation:\\s*".{1,150}"\\s*(ok|okay|sure|alright|fine),?\\s+i (will|can) do it( tomorrow| later)?\\.?$', '^based on the quotation:\\s*".{1,150}"\\s*i (will|can) do it( tomorrow| later)?\\.?$', '^based on the quotation:\\s*".{1,150}"\\s*i\\\'ll do it( tomorrow| later)?\\.?$', '^waiting for this.*$', '^waiting for .* to be merged.*$', '^this is ready for review\\.?$', '^ready to merge\\.?$', '^ready for merge\\.?$', '^ready to go\\.?$', '^this is good to go\\.?$', '^with .* merged, i think this is ready\\.?$', '^i think this is ready\\.?$', '^closed\\.$', '^closed this pr\\.?$', '^closing this issue\\.?$', '^closed in favor of.*$', '^reopening\\.?$', "^i'll close this pr\\.?$", '^i will close this pr\\.?$', '^thanks for the update\\.?$', '^thanks for the information\\.?$', '^thanks for the review\\.?$', '^thanks for the feedback\\.?$', '^thanks for checking\\.?$', '^thanks for the clarification\\.?$', '^thank you for your guidance\\.?$', '^thank you for the review\\.?$', '^thank you for your feedback\\.?$', '^thank you for your comments\\.?$', '^thanks all\\.?$', '^thx all\\.?$', '^manually run the test job again\\.?$', '^re-running ci\\.?$', '^rerunning ci\\.?$', '^ci errors look legit\\.?$', '^(the )?ci is green now!?$', '^test link:.*$', '^passed in \\d+.*$', '^passed again.*$', '^can the pipeline be triggered again\\??.*$', '^can the pipeline be re-triggered\\??.*$', '^could the pipeline be triggered again\\??.*$', '^could the pipeline be re-triggered\\??.*$', '^please trigger the pipeline again\\.?.*$', '^please re-trigger the pipeline\\.?.*$', '^please rerun the pipeline\\.?.*$', '^please rerun ci\\.?.*$', '^please rerun the ci\\.?.*$', '^can ci be triggered again\\??.*$', '^could ci be triggered again\\??.*$', '^can the checks be triggered again\\??.*$', '^could the checks be triggered again\\??.*$', '^please rerun the checks\\.?.*$', '^checks are going now\\.?$']

def normalize_text(texto):
    return str(texto).strip().lower()

def matches_any_fullmatch(texto, patterns):
    return any((re.fullmatch(pattern, texto, flags=re.DOTALL) for pattern in patterns))

def matches_any_search(texto, patterns):
    return any((re.search(pattern, texto, flags=re.DOTALL) for pattern in patterns))

def starts_with_automated_marker(texto):
    texto = texto.strip().lower()
    return any((texto.startswith(marker) for marker in AUTOMATED_PREFIX_MARKERS))

def is_only_slash_commands(texto):
    lineas = [line.strip() for line in texto.splitlines() if line.strip()]
    if not lineas:
        return False
    return all((re.fullmatch('/[a-zA-Z0-9_-]+(\\s+[a-zA-Z0-9_.-]+)?', line) for line in lineas))

def remove_low_value_tokens(texto):
    texto_limpio = texto
    for placeholder in PLACEHOLDERS_TO_REMOVE:
        texto_limpio = texto_limpio.replace(placeholder, ' ')
    for pattern in GENERIC_USER_PLACEHOLDERS:
        texto_limpio = re.sub(pattern, ' ', texto_limpio)
    for pattern in LOW_VALUE_VISUAL_PATTERNS:
        texto_limpio = re.sub(pattern, ' ', texto_limpio)
    texto_limpio = re.sub('\\s+', ' ', texto_limpio).strip()
    return texto_limpio

def es_hard_noise(texto):
    texto = normalize_text(texto)
    if starts_with_automated_marker(texto):
        return True
    if matches_any_fullmatch(texto, HARD_EXACT_PATTERNS):
        return True
    if is_only_slash_commands(texto):
        return True
    if matches_any_search(texto, HARD_PHRASE_PATTERNS):
        return True
    texto_limpio = remove_low_value_tokens(texto)
    if len(texto_limpio) <= 3:
        return True
    return False

def es_operational_noise(texto):
    texto = normalize_text(texto)
    if es_hard_noise(texto):
        return False
    if matches_any_fullmatch(texto, OPERATIONAL_EXACT_PATTERNS):
        return True
    if matches_any_fullmatch(texto, OPERATIONAL_PHRASE_PATTERNS):
        return True
    return False