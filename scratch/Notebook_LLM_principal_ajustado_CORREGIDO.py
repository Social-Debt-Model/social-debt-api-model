# paso 1
# ============================================
# INSTALAR DEPENDENCIAS
# ============================================
!pip install -q openai pandas openpyxl scikit-learn


# ---

# ============================================
# API KEY EN COLAB SECRETS O VARIABLE DE ENTORNO
# ============================================
# OPCIÓN RECOMENDADA EN COLAB:
# 1. Abre el panel izquierdo de Colab.
# 2. Entra a Secrets / icono de llave.
# 3. Crea un secreto con este nombre exacto: OPENAI_API_KEY
# 4. Activa "Notebook access".
#
# NO pegues aquí tu API key real si vas a compartir el notebook.
# Si necesitas una prueba temporal, usa una celda privada y no la subas a GitHub:
# import os
# os.environ["OPENAI_API_KEY"] = "sk-..."


# ---

# paso 2

# ============================================
# CONFIGURACIÓN SEGURA DE API KEY
# ============================================
# Usa primero Colab Secrets y, si no existe, usa variable de entorno OPENAI_API_KEY.

import os
from openai import OpenAI

api_key = None

try:
    from google.colab import userdata
    api_key = userdata.get("OPENAI_API_KEY")
except Exception:
    api_key = None

if not api_key:
    api_key = os.getenv("OPENAI_API_KEY")

if not api_key:
    raise ValueError(
        "No se encontró OPENAI_API_KEY. Configúrala en Colab Secrets o como variable de entorno."
    )

client = OpenAI(api_key=api_key.strip())
print("Cliente OpenAI configurado correctamente.")


# ---


# paso 3
# ============================================
# IMPORTS Y CONFIGURACIÓN GENERAL
# ============================================
import re
import time
import pandas as pd
from pathlib import Path
from sklearn.metrics import accuracy_score, classification_report, confusion_matrix

# Archivo de entrada
INPUT_FILE = "dataset_limpio_final.xlsx"

# Archivo principal de salida
OUTPUT_FILE = "dataset_mapeo.xlsx"

TEXT_COLUMN = "comment_body_clean_final"
GOLD_COLUMN = "Normalized_cause"

# =====================================================
# COLUMNAS DE PREDICCIÓN FINAL
# =====================================================

OUTPUT_COLUMN_CODE = "predicted_code"
OUTPUT_COLUMN_LABEL = "predicted_label"
OUTPUT_COLUMN_CONF = "Confidence"
OUTPUT_COLUMN_REVIEW = "Requires_revision"
OUTPUT_COLUMN_STATUS = "State_prediction"

OUTPUT_COLUMN_CODE_LLM = "Code_cause_llm"
OUTPUT_COLUMN_LABEL_LLM = "predicted_label_llm"
OUTPUT_COLUMN_CONF_LLM = "Confidence_llm"
OUTPUT_COLUMN_RULE = "Priority_rule_applied"

FINAL_CAUSE_CODE_COLUMN = "final_cause_code"
FINAL_CAUSE_COLUMN = "final_cause_for_analysis"

MODEL = "gpt-4o-mini"
PROCESS_ALL = True
SAMPLE_SIZE = 100
SLEEP_SECONDS = 0.2
CONFIDENCE_THRESHOLD = 0.70

# En esta versión, el LLM es el clasificador principal.
USE_LLM_AS_MAIN_CLASSIFIER = True


# ---

# paso 4
# ============================================
# CATÁLOGO FINAL DE 8 CAUSAS
# ============================================
CAUSES = {
    "A": "Communication and shared understanding breakdowns",
    "B": "Coordination and workflow misalignment",
    "C": "Technical complexity, compatibility, and system constraints",
    "D": "Organizational and procedural workflow constraints",
    "E": "Collaboration and interpersonal tensions",
    "F": "Knowledge, documentation, and standards deficiencies",
    "G": "Resource, tooling, access, and validation dependencies",
    "H": "No identifiable sociotechnical cause",
}


# ============================================
# TÉRMINOS PARA DETECTAR COMPORTAMIENTO TÉCNICO
# ============================================

TECHNICAL_BEHAVIOR_TERMS = [
    "diagnostic",
    "diagnostics",
    "editor",
    "vscode",
    "vs code",
    "runtime",
    "configuration",
    "config",
    "compatibility",
    "system behavior",
    "extension",
    "extensions",
    "tsserver",
    "marker",
    "markers",
    "uri",
    "api",
]

VALID_CODES = set(CAUSES.keys())
UNKNOWN_CODE = "H"

LABEL_TO_CODE = {
    label.lower().strip(): code
    for code, label in CAUSES.items()
}

def _norm_label(value):
    if pd.isna(value):
        return ""
    value = str(value).strip()
    value = re.sub(r"\s+", " ", value)
    return value

def label_to_code(label):
    """Convierte código o etiqueta textual final a código A-H."""
    label = _norm_label(label)

    if not label:
        return None

    upper = label.upper()

    if upper in VALID_CODES:
        return upper

    key = label.lower()

    if key in LABEL_TO_CODE:
        return LABEL_TO_CODE[key]

    return None

def code_to_label(code):
    """Convierte código A-H a nombre completo final."""
    code = label_to_code(code)

    if code in VALID_CODES:
        return CAUSES[code]

    return None

def normalize_cause_name(value):
    """Devuelve nombre final de causa desde código o nombre actual."""
    code = label_to_code(value)

    if code in VALID_CODES:
        return CAUSES[code]

    return None

print("Códigos válidos:", sorted(VALID_CODES))
print("Total causas:", len(CAUSES))
print("UNKNOWN_CODE:", UNKNOWN_CODE)



# ---

# paso 5
# ============================================
# PASO 5. PROMPT DEL LLM — CLASIFICADOR PRINCIPAL
# ============================================

SYSTEM_PROMPT = f"""
You are an expert in software engineering, specializing in sociotechnical factors and social debt in software development teams.

Your role is to analyze GitHub comments and identify the underlying cause of sociotechnical issues based only on observable evidence in the text.

The LLM is the main classifier. Deterministic rules may later be used only as conservative post-processing.

=================
TASK
=================
Read the GitHub comment and classify it by identifying the underlying cause.
Select exactly ONE cause from the reduced label catalog.
Use the signals only as guidance. Do not require all signals to be present.
Use quoted or referenced context only when it helps interpret the author's current message.
Do not classify a quoted fragment alone if the author's own text contains enough evidence.

================================
REDUCED LABEL CATALOG — 8 CAUSES
================================

A: Communication and shared understanding breakdowns
Use when the comment shows unclear meaning, ambiguous wording, misunderstanding, confusion, lack of clarification, incorrect interpretation, flawed reasoning, or lack of shared understanding.
Typical signals: "I misunderstood", "I get you now", "not sure what you mean", "unclear", "confusing", "clarify", "correction to my analysis".

B: Coordination and workflow misalignment
Use when the comment concerns coordination between contributors, reviewers, PRs, or tasks.
This includes review synchronization, split PRs, separate PRs, follow-up PRs, merge order, re-approval after changes, availability to continue work, deciding next steps, handoff, consensus building, applying reviewer comments, or coordinating who should do what.
Do NOT select B merely because a normal review is mentioned. Select B when the comment reflects coordination of work, sequencing, responsibilities, or contributor synchronization.

C: Technical complexity, compatibility, and system constraints
Use when the comment primarily discusses technical behavior, implementation logic, API behavior, configuration, compatibility, architecture, dependencies, runtime behavior, system constraints, feature gates, context handling, syntax rules, package behavior, version-specific behavior, database-specific behavior, or preserving previous system behavior.
This category is about the technical nature of the software problem itself.

D: Organizational and procedural workflow constraints
Use when the comment concerns formal project procedures or administrative workflow constraints.
This includes release branches, future releases, milestones, labels, ticket flags, approval process, merge permissions, triage state, patch status, release process, upstream adoption, or formal project requirements.
Do NOT select D for simple automated bot messages, routine status updates, or CI/test results unless the comment explicitly describes a formal procedural blockage.

E: Collaboration and interpersonal tensions
Use when the comment expresses interpersonal tension, perceived unfairness, dismissive interaction, low trust, frustration caused by people, negative tone, conflict between contributors, or harmful collaboration climate.
If the same comment mainly describes conflicting review directions or lack of synchronization, prefer B.

F: Knowledge, documentation, and standards deficiencies
Use when the comment concerns documentation, tutorials, PR description, release notes, deprecation notes, versionchanged notes, missing explanation, unclear instructions, official documentation quality, coding standards, contribution guidance, or knowledge transfer.
Prefer F when the main focus is how something should be documented, explained, named, written, or standardized.

G: Resource, tooling, access, and validation dependencies
Use when the comment concerns tests, CI, dashboards, perf dash, coverage, logs, monitoring, linter, gofmt, compile errors, minimal reproduction, local tests, external tests, validation jobs, clusters, repository access, permissions, tooling limitations, maintainer dependency, resource availability, or needing validation to proceed.
This category includes validation and tooling evidence even if the comment does not explicitly say "blocked".

H: No identifiable sociotechnical cause
Use only when the comment is mainly neutral, minimal, or informational and contains no clear sociotechnical evidence.
Typical cases: simple thanks, LGTM, ping, done, related issue link, closes link, upload notice, simple acknowledgement, or routine status update without unresolved technical, validation, documentation, coordination, communication, procedural, or interpersonal evidence.

=================
PRIORITY GUIDELINES
=================

1. Do NOT select H if the comment contains clear evidence of technical behavior, compatibility, configuration, tests, CI, dashboard, logs, monitoring, linter, compile error, documentation, misunderstanding, coordination, procedural constraint, access limitation, or interpersonal tension.

2. Select G when the comment mentions validation or tooling evidence such as tests, test cases, CI, dashboard, perf dash, coverage, logs, monitoring, linter, gofmt, compile error, minimal reproduction, external tests, local tests, validation jobs, clusters, or tooling used to verify behavior.

3. Select F when the main focus concerns documentation, PR description, release notes, deprecation notes, tutorials, wording, instructions, standards, or making documentation accurate and easy to understand.

4. Select A when the main issue is unclear meaning, ambiguity, misunderstanding, correction of a wrong interpretation, lack of clarification, confusing wording, or lack of shared understanding.

5. Select B when the comment focuses on coordinating work between people or PRs: review again, split PRs, separate PRs, follow-up PR, merge order, re-approval, availability, consensus, next steps, handoff, applying reviewer comments, or synchronizing contributors.

6. Select C when the comment focuses on software behavior, implementation details, API behavior, compatibility, configuration, runtime behavior, architecture, dependencies, feature gates, context cancellation, selectors, syntax, packages, or preserving previous behavior.

7. Select D when the comment focuses on formal workflow or project procedure: release branches, future releases, milestones, labels, ticket flags, approval process, merge permission, triage, patch status, release process, or upstream adoption.

8. Select E only when there is explicit interpersonal tension, unfairness, dismissive interaction, frustration caused by people, low trust, or negative collaboration climate.

9. Routine review requests such as "can you review?", "PTAL", "ping", "take a look", or simple mentions should be H unless the comment also contains clear coordination, technical, validation, documentation, or procedural evidence.

10. Simple status updates such as "done", "rebased", "fixed", "updated", "thanks", or "applied suggestions" should be H unless the comment explains what was technically changed, documented, validated, coordinated, or procedurally required.

11. If a comment contains both documentation and technical content, choose F when the main request is to document, explain, clarify wording, update release notes, or improve official guidance. Choose C when the main issue is the technical behavior itself.

12. If a comment contains both workflow and validation evidence, choose G when the main focus is test/CI/tooling/validation. Choose D when the main focus is formal approval, release process, merge permission, labels, milestones, or ticket status.

13. If a comment contains both coordination and procedural evidence, choose B when the main focus is people coordinating work. Choose D when the main focus is a formal project rule or administrative requirement.

=================
EXAMPLES
=================

Comment: "Thank you for this PR! Code looks good, did you run the perf dash to check it works as intended?"
Answer: G|0.90

Comment: "External tests ran successfully."
Answer: G|0.90

Comment: "Here's a minimal reproduction."
Answer: G|0.85

Comment: "Sorry, I misunderstood your comment. I've made the correction."
Answer: A|0.90

Comment: "I gotcha, updated the PR description."
Answer: F|0.85

Comment: "Please edit the PR body text to include the release note."
Answer: F|0.90

Comment: "I split this into small PRs for convenient review and better control merge order."
Answer: B|0.90

Comment: "What happens next? You all approved the changes and I applied the comments."
Answer: B|0.85

Comment: "This was added to preserve the previous behavior and avoid changing API semantics."
Answer: C|0.90

Comment: "If the secret exists, load it from configuration; otherwise fall back to the text file."
Answer: C|0.85

Comment: "Could this be added to the 5.2 branch in a future release?"
Answer: D|0.85

Comment: "Kubernetes was released today. Now EKS has to pick this up."
Answer: D|0.85

Comment: "The interaction in this discussion feels dismissive and unfair."
Answer: E|0.90

Comment: "Can you review?"
Answer: H|0.90

Comment: "Ping a GitHub user."
Answer: H|0.95

Comment: "Related issue: /url_reference"
Answer: H|0.95

Comment: "LGTM."
Answer: H|0.95

Comment: "Done."
Answer: H|0.95

=================
OUTPUT FORMAT
=================
Return ONLY this format:

CODE|CONFIDENCE

Where:
- CODE is one valid code from A to H.
- CONFIDENCE is a decimal number between 0 and 1.
- Do not return the cause text.
- Do not return JSON.
- Do not explain your answer.
"""


# ---

# paso 6
# ============================================
# FUNCIONES DE LIMPIEZA, REGLAS Y CLASIFICACIÓN
# LLM PRINCIPAL + REGLAS COMO POSPROCESAMIENTO CONSERVADOR
# ============================================

def clean_code(output):
    """Extrae un código válido A-H desde la salida del modelo."""
    if output is None:
        return UNKNOWN_CODE

    text = str(output).strip().upper().replace('"', '').replace("'", "").strip()

    first_part = text.split("|")[0].strip()
    if first_part in VALID_CODES:
        return first_part

    for code in sorted(VALID_CODES):
        if re.search(rf"(?<![A-Z]){re.escape(code)}(?![A-Z])", text):
            return code

    return UNKNOWN_CODE


def clean_confidence(output, default=0.50):
    """Extrae una confianza entre 0 y 1 desde la salida del modelo."""
    if output is None:
        return 0.0

    text = str(output).strip().replace(",", ".")
    matches = re.findall(r"(?<!\d)(?:0(?:\.\d+)?|1(?:\.0+)?)(?!\d)", text)

    if matches:
        try:
            value = float(matches[-1])
            return max(0.0, min(1.0, value))
        except ValueError:
            pass

    return float(default)


def parse_model_output(output):
    """Devuelve (code, confidence) desde la respuesta del modelo."""
    return clean_code(output), clean_confidence(output, default=0.50)


def move_column_after(df, column_to_move, after_column):
    """Reordena columnas para que una columna quede justo después de otra."""
    if column_to_move not in df.columns or after_column not in df.columns:
        return df

    cols = list(df.columns)
    cols.remove(column_to_move)
    insert_at = cols.index(after_column) + 1
    cols.insert(insert_at, column_to_move)

    return df[cols]


def prepare_comment_for_llm(comment):
    """Prepara el comentario para el LLM."""
    if pd.isna(comment):
        return ""

    text = str(comment)
    text = re.sub(r"\n{3,}", "\n\n", text).strip()

    return text


def contains_any(text, patterns):
    """Verifica si algún patrón aparece en el texto."""
    return any(
        re.search(pattern, str(text), flags=re.IGNORECASE)
        for pattern in patterns
    )


# ============================================
# PATRONES DE ALTA PRECISIÓN
# ============================================

VALIDATION_EVIDENCE_PATTERNS = [
    r"\btests?\b",
    r"\bci\b",
    r"continuous integration",
    r"perf dash",
    r"dashboard",
    r"coverage",
    r"\blogs?\b",
    r"monitoring",
    r"linter",
    r"gofmt",
    r"compile error",
    r"build error",
    r"minimal reproduction",
    r"reproduction",
    r"external tests?",
    r"local tests?",
    r"validation job",
    r"cluster",
    r"job failed",
    r"job failing",
    r"ran successfully",
    r"got a green",
    r"green build",
    r"red build",
]

DOCUMENTATION_EVIDENCE_PATTERNS = [
    r"documentation",
    r"\bdocs\b",
    r"tutorial",
    r"pr description",
    r"pull request description",
    r"release notes?",
    r"release-note",
    r"versionchanged",
    r"deprecation notes?",
    r"deprecation message",
    r"document .* behavior",
    r"document .* decision",
    r"wording",
    r"instructions",
    r"standards?",
    r"easy to understand",
    r"accurate .* document",
]

COMMUNICATION_EVIDENCE_PATTERNS = [
    r"misunderstood",
    r"i misunderstood",
    r"i get you now",
    r"not sure what you mean",
    r"what do you mean",
    r"unclear",
    r"confusing",
    r"ambiguous",
    r"clarif(?:y|ication)",
    r"correct(?:ed|ion) .* analysis",
    r"flawed reasoning",
    r"correct my analysis",
    r"reasoning .* flawed",
    r"i see now",
    r"i genuinely appreciate .* correct",
]

COORDINATION_EVIDENCE_PATTERNS = [
    r"split .* prs?",
    r"smaller prs?",
    r"separate pr",
    r"follow[- ]?up pr",
    r"merge order",
    r"review again",
    r"re-?approve",
    r"reapproval",
    r"what happens next",
    r"next steps",
    r"availability",
    r"available .* continue",
    r"consensus",
    r"same page",
    r"handoff",
    r"synchroni[sz]e",
    r"squashed .* commits?",
    r"force[- ]?pushed",
    r"resolved .* conflict",
    r"apply(?:ing|ied) .* comments",
]

PROCEDURAL_EVIDENCE_PATTERNS = [
    r"release branch",
    r"future release",
    r"milestones?",
    r"labels?",
    r"ticket flags?",
    r"patch needs improvement",
    r"needs tests",
    r"needs rebase",
    r"approval process",
    r"merge permission",
    r"triage",
    r"release process",
    r"upstream adoption",
    r"eks .* pick this up",
    r"todo list",
    r"critical .* todo",
]

INTERPERSONAL_EVIDENCE_PATTERNS = [
    r"unfair",
    r"dismissive",
    r"ignored",
    r"not being heard",
    r"low trust",
    r"toxic",
    r"code of conduct",
    r"interpersonal tension",
    r"frustrat(?:ed|ion) .* review",
]

TECHNICAL_EVIDENCE_PATTERNS = [
    r"\bconfiguration\b",
    r"\bconfig\b",
    r"\bcompatibility\b",
    r"\bimplementation\b",
    r"\bruntime\b",
    r"\bfeature gate\b",
    r"\bcontext cancellation\b",
    r"\bsecret\b",
    r"\blogger\b",
    r"\bselector\b",
    r"\bsyntax\b",
    r"\bpackage\b",
    r"\bclient-go\b",
    r"\bdefault behavior\b",
    r"\bprevious behavior\b",
    r"\bpreserve .* behavior\b",
    r"\bsystem behavior\b",
    r"\bdependency\b",
    r"\bdependencies\b",
    r"\barchitecture\b",
    r"\bedge case\b",
    r"\bsqlite\b",
    r"\bmysql\b",
    r"\boracle\b",
    r"\bpython 3\.10\b",
]

MINIMAL_H_PATTERNS = [
    r"^\s*(thanks|thank you|thx|ty|lgtm|done|fixed|ping|ptal)\s*[.!:]*\s*$",
    r"^\s*related issue\s*:?\s*/?url_reference\s*$",
    r"^\s*closes\s+/?url_reference\s*$",
    r"^\s*a github user is mentioned\.?\s*$",
    r"^\s*/(?:lgtm|approve|unhold|hold|retest|test|cc|assign|unassign)\s*$",
]

BOT_TEMPLATE_PATTERNS = [
    r"automated_bot_message",
    r"workflow_ci",
    r"ci_status_reference",
    r"flaky tests guide",
    r"retesting failed pr",
    r"pull-request has been approved",
    r"lgtm label has been added",
    r"this issue is currently awaiting triage",
]

SOCIAL_OR_MINIMAL_CONTEXT_PATTERNS = [
    r"\blgtm\b",
    r"\blooks good\b",
    r"\bsounds good\b",
    r"\bthank you\b",
    r"\bthanks\b",
    r"\bappreciate\b",
    r"\bnice work\b",
    r"\bgreat job\b",
    r"\bkeen to see\b",
    r"\bfine to me\b",
    r"\bgot it\b",
]

TECHNICAL_PROBLEM_PATTERNS = [
    r"\berror\b",
    r"\bfail(?:ed|ing)?\b",
    r"\bbroken\b",
    r"\bbug\b",
    r"\bissue\b",
    r"\bproblem\b",
    r"\bconstraint\b",
    r"\bincompatible\b",
    r"\bregression\b",
    r"\bdoes not work\b",
    r"\bcannot\b",
    r"\bcan't\b",
]

VALIDATION_PROBLEM_PATTERNS = [
    r"\bcannot\b",
    r"\bcan't\b",
    r"\bfail(?:ed|ing)?\b",
    r"\bflake\b",
    r"\bflaky\b",
    r"\bretry\b",
    r"\btimeout\b",
    r"\btrigger\b",
    r"\bmissing\b",
    r"\bred\b",
    r"\bgreen ci\b",
    r"\badded test case\b",
    r"\brun\b",
    r"\bran\b",
]


def get_rule_candidate(comment):
    """
    Devuelve una regla de alta precisión basada en evidencia observable.
    Esta regla NO reemplaza al LLM siempre; apply_priority_rules decide cuándo usarla.
    """
    text = re.sub(r"\s+", " ", str(comment or "")).strip()

    has_evidence = any([
        contains_any(text, VALIDATION_EVIDENCE_PATTERNS),
        contains_any(text, DOCUMENTATION_EVIDENCE_PATTERNS),
        contains_any(text, COMMUNICATION_EVIDENCE_PATTERNS),
        contains_any(text, COORDINATION_EVIDENCE_PATTERNS),
        contains_any(text, PROCEDURAL_EVIDENCE_PATTERNS),
        contains_any(text, INTERPERSONAL_EVIDENCE_PATTERNS),
        contains_any(text, TECHNICAL_EVIDENCE_PATTERNS),
    ])

    if contains_any(text, BOT_TEMPLATE_PATTERNS) and not has_evidence:
        return "H", 0.95, "H_bot_template_without_evidence"

    if contains_any(text, MINIMAL_H_PATTERNS) and not has_evidence:
        return "H", 0.95, "H_minimal_without_evidence"

    if contains_any(text, COMMUNICATION_EVIDENCE_PATTERNS):
        return "A", 0.90, "A_communication_evidence"

    if contains_any(text, DOCUMENTATION_EVIDENCE_PATTERNS):
        return "F", 0.88, "F_documentation_evidence"

    if contains_any(text, VALIDATION_EVIDENCE_PATTERNS):
        return "G", 0.88, "G_validation_tooling_evidence"

    if contains_any(text, INTERPERSONAL_EVIDENCE_PATTERNS):
        return "E", 0.88, "E_interpersonal_tension_evidence"

    if contains_any(text, PROCEDURAL_EVIDENCE_PATTERNS):
        return "D", 0.85, "D_procedural_workflow_evidence"

    if contains_any(text, COORDINATION_EVIDENCE_PATTERNS):
        return "B", 0.85, "B_coordination_workflow_evidence"

    if contains_any(text, TECHNICAL_EVIDENCE_PATTERNS):
        return "C", 0.82, "C_technical_evidence"

    return None, None, "none"


def strong_override_allowed(comment, rule_code):
    """
    Evita que reglas débiles cambien H hacia otra causa.
    """
    text = re.sub(r"\s+", " ", str(comment or "")).lower().strip()

    if contains_any(text, SOCIAL_OR_MINIMAL_CONTEXT_PATTERNS):
        if rule_code == "G":
            return (
                contains_any(text, VALIDATION_EVIDENCE_PATTERNS)
                and contains_any(text, VALIDATION_PROBLEM_PATTERNS)
            )
        if rule_code == "C":
            return (
                contains_any(text, TECHNICAL_EVIDENCE_PATTERNS)
                and contains_any(text, TECHNICAL_PROBLEM_PATTERNS)
            )
        if rule_code == "F":
            return contains_any(text, DOCUMENTATION_EVIDENCE_PATTERNS)

        return False

    if rule_code == "C":
        return (
            contains_any(text, TECHNICAL_EVIDENCE_PATTERNS)
            and contains_any(text, TECHNICAL_PROBLEM_PATTERNS)
        )

    if rule_code == "G":
        return (
            contains_any(text, VALIDATION_EVIDENCE_PATTERNS)
            and contains_any(text, VALIDATION_PROBLEM_PATTERNS)
        )

    if rule_code == "F":
        return contains_any(text, DOCUMENTATION_EVIDENCE_PATTERNS)

    return True


def apply_priority_rules(comment, llm_code, llm_conf):
    """
    Posprocesamiento conservador.
    """

    llm_code = label_to_code(llm_code) or UNKNOWN_CODE
    llm_conf = float(llm_conf or 0.0)

    rule_code, rule_conf, rule_name = get_rule_candidate(comment)

    if llm_code not in VALID_CODES:
        if rule_code in VALID_CODES:
            return rule_code, rule_conf, rule_name
        return UNKNOWN_CODE, 0.50, "invalid_llm_code_fallback"

    if rule_code is None:
        return llm_code, llm_conf, "none"

    if rule_code == "H" and rule_name.startswith("H_"):
        return "H", max(llm_conf, rule_conf), rule_name

    if llm_code == "H" and rule_code != "H":
        if strong_override_allowed(comment, rule_code):
            return rule_code, max(llm_conf, rule_conf), rule_name + "_override_H"
        return llm_code, llm_conf, "override_rejected_keep_H"

    if llm_conf < CONFIDENCE_THRESHOLD and rule_code != llm_code:
        return rule_code, max(llm_conf, rule_conf), rule_name + "_low_conf_override"

    if rule_code == llm_code:
        return llm_code, max(llm_conf, rule_conf), rule_name + "_confirmed"

    return llm_code, llm_conf, "none_llm_kept"


def classify(comment):
    """
    Clasifica un comentario con LLM como clasificador principal.

    Devuelve:
    final_code, final_label, final_conf, needs_review, status,
    llm_code, llm_label, llm_conf, rule_applied.
    """

    prepared_comment = prepare_comment_for_llm(comment)

    if not prepared_comment:
        return (
            UNKNOWN_CODE,
            CAUSES[UNKNOWN_CODE],
            0.0,
            True,
            "EMPTY_TEXT",
            UNKNOWN_CODE,
            CAUSES[UNKNOWN_CODE],
            0.0,
            "empty",
        )

    if "client" not in globals() or client is None:
        raise NameError("client no está definido. Ejecuta primero la celda de configuración de API key.")

    user_prompt = f"""Classify the following GitHub comment.

Comment:
{prepared_comment}
"""

    try:
        response = client.chat.completions.create(
            model=MODEL,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": user_prompt},
            ],
            max_completion_tokens=10,
            temperature=0,
        )

        raw_output = response.choices[0].message.content

        llm_code, llm_conf = parse_model_output(raw_output)
        llm_code = label_to_code(llm_code) or UNKNOWN_CODE

        final_code, final_conf, rule_applied = apply_priority_rules(
            prepared_comment,
            llm_code,
            llm_conf,
        )

        final_code = label_to_code(final_code) or UNKNOWN_CODE
        final_label = CAUSES.get(final_code, CAUSES[UNKNOWN_CODE])
        llm_label = CAUSES.get(llm_code, CAUSES[UNKNOWN_CODE])

        needs_review = float(final_conf) < CONFIDENCE_THRESHOLD

        if rule_applied.startswith("none"):
            status = "REVIEW" if needs_review else "ACCEPTED"
        elif "override" in rule_applied:
            status = "RULE_OVERRIDE"
        elif "confirmed" in rule_applied:
            status = "RULE_CONFIRMED"
        else:
            status = "RULE_APPLIED"

        return (
            final_code,
            final_label,
            float(final_conf),
            bool(needs_review),
            status,
            llm_code,
            llm_label,
            float(llm_conf),
            rule_applied,
        )

    except Exception as e:
        print(f"Error al clasificar comentario: {e}")

        return (
            UNKNOWN_CODE,
            CAUSES[UNKNOWN_CODE],
            0.0,
            True,
            "API_ERROR",
            UNKNOWN_CODE,
            CAUSES[UNKNOWN_CODE],
            0.0,
            "api_error",
        )

# ---


# paso 7
# ============================================
# VALIDACIÓN RÁPIDA DE REGLAS SIN LLAMAR A LA API
# ============================================
# Esta prueba verifica solo el posprocesamiento conservador.
# No reemplaza la prueba principal con LLM.

test_comments_rules = [
    "Can you review?",
    "Ping a GitHub user",
    "Thank you for this PR! Code looks good, did you run the perf dash to check it works as intended?",
    "External tests ran successfully: /url_reference",
    "Sorry, I misunderstood your comment. I've made the correction.",
    "I gotcha a GitHub user, updated the pr description",
    "Resolved trival conflict to another feature gate. Please review again, thanks.",
    "Here's a minimal reproduction /url_reference",
]

for c in test_comments_rules:
    rule_code, rule_conf, rule_name = get_rule_candidate(c)
    final_code, final_conf, rule_applied = apply_priority_rules(c, "H", 0.50)
    print("Comentario:", c[:140].replace("\n", " "))
    print("Regla candidata:", rule_code, rule_name)
    print("Si LLM=H, código final:", final_code)
    print("Causa final:", CAUSES[final_code])
    print("Confianza:", final_conf)
    print("Regla aplicada:", rule_applied)
    print("-" * 100)


# ---

# paso 8
# ============================================
# CARGAR DATASET
# ============================================
input_path = Path(INPUT_FILE)

if not input_path.exists():
    candidate_names = [
        INPUT_FILE,
        "dataset_limpio_final.xlsx",

    ]
    candidate_dirs = [Path.cwd(), Path("/content"), Path("/mnt/data")]
    candidates = []
    for base in candidate_dirs:
        for name in candidate_names:
            p = base / name
            if p not in candidates:
                candidates.append(p)
    input_path = next((p for p in candidates if p.exists()), None)

if input_path is None or not input_path.exists():
    raise FileNotFoundError(f"No se encontró el archivo de entrada: {INPUT_FILE}")

df = pd.read_excel(input_path)
print(f"Dataset cargado desde: {input_path}")

if TEXT_COLUMN not in df.columns:
    if "comment_body_raw" in df.columns:
        print(f"Advertencia: no existe '{TEXT_COLUMN}'. Se usará 'comment_body_raw'.")
        TEXT_COLUMN = "comment_body_raw"
    else:
        raise ValueError(f"No existe columna de texto. Columnas disponibles: {list(df.columns)}")

print("Columnas iniciales:", df.columns.tolist())
print("Filas originales:", len(df))

COLUMNS_TO_REBUILD = [
    OUTPUT_COLUMN_CODE,
    OUTPUT_COLUMN_LABEL,
    OUTPUT_COLUMN_CONF,
    OUTPUT_COLUMN_REVIEW,
    OUTPUT_COLUMN_STATUS,
    OUTPUT_COLUMN_CODE_LLM,
    OUTPUT_COLUMN_LABEL_LLM,
    OUTPUT_COLUMN_CONF_LLM,
    OUTPUT_COLUMN_RULE,
    FINAL_CAUSE_CODE_COLUMN,
    FINAL_CAUSE_COLUMN,
    "true_code",
    "evaluable",
    "match",
    "predicted_cause_name",
    "final_label",
]

existing_rebuild_cols = [c for c in COLUMNS_TO_REBUILD if c in df.columns]
if existing_rebuild_cols:
    print("Se eliminarán columnas de predicción anteriores para evitar mezcla:", existing_rebuild_cols)
    df = df.drop(columns=existing_rebuild_cols)

df = df.dropna(subset=[TEXT_COLUMN]).copy()
df[TEXT_COLUMN] = df[TEXT_COLUMN].astype(str)
df = df[df[TEXT_COLUMN].str.strip() != ""].copy()
df = df.reset_index(drop=True)

print("Filas con texto válido:", len(df))

# ============================================
# MODO PRUEBA - SOLO 10 REGISTROS
# ============================================
TEST_MODE = False

if TEST_MODE:
    df = df.head(20).copy()
    df = df.reset_index(drop=True)
    print("Modo prueba activado. Filas usadas:", len(df))


# ---

# paso 9
# ============================================
# SELECCIONAR DATOS A PROCESAR
# ============================================
if not PROCESS_ALL:
    sample_n = min(SAMPLE_SIZE, len(df))
    df = df.head(sample_n).copy()
    print("Modo prueba activado.")
    print(f"Procesando {sample_n} registros.")
else:
    print("Procesando dataset completo.")
    print(f"Total registros: {len(df)}")


# ---

# paso 10
# ============================================
# PRUEBA CORTA DEL LLM ANTES DE PROCESAR TODO
# ============================================
# Ejecuta esta celda antes de clasificar el dataset completo.
# Si aparece API_ERROR, no continúes hasta corregir la API key.

quick_tests = [
    "Thank you for this PR! Code looks good, did you run the perf dash to check it works as intended?",
    "External tests ran successfully: /url_reference",
    "Sorry, I misunderstood your comment. I've made the correction.",
]

for t in quick_tests:
    result = classify(t)
    print("\nComentario:", t)
    print("FINAL:", result[0], result[1])
    print("CONF:", result[2])
    print("STATUS:", result[4])
    print("LLM:", result[5], result[6], result[7])
    print("RULE:", result[8])


# ---

# paso 11
# ============================================================
# CLASIFICAR DATASET Y EXPORTAR COLUMNAS FINALES
# ============================================================

# Normalized_cause NO se sobrescribe.
# Se regeneran desde cero Code_cause y predicted_label.
# Code_cause = código A-H.
# predicted_label = nombre completo de causa.

# -----------------------------
# 1. Definir columnas de salida
# -----------------------------

output_cols = [
    OUTPUT_COLUMN_CODE,
    OUTPUT_COLUMN_LABEL,
    OUTPUT_COLUMN_CONF,
    OUTPUT_COLUMN_REVIEW,
    OUTPUT_COLUMN_STATUS,
    OUTPUT_COLUMN_CODE_LLM,
    OUTPUT_COLUMN_LABEL_LLM,
    OUTPUT_COLUMN_CONF_LLM,
    OUTPUT_COLUMN_RULE,
]

for col in output_cols:
    df[col] = None

# -----------------------------
# 2. Clasificar registros
# -----------------------------



for n, (idx, row) in enumerate(df.iterrows(), start=1):

    text = row.get(TEXT_COLUMN, "")

    (
        code,
        label,
        conf,
        needs_review,
        status,
        llm_code,
        llm_label,
        llm_conf,
        rule_applied,
    ) = classify(text)

    # Normalizar códigos
    code = label_to_code(code) or UNKNOWN_CODE
    llm_code = label_to_code(llm_code) or UNKNOWN_CODE

    # Validar contra catálogo A-H
    if code not in VALID_CODES:
        code = UNKNOWN_CODE

    if llm_code not in VALID_CODES:
        llm_code = UNKNOWN_CODE

    # Mapear nombres completos
    label = CAUSES.get(code, CAUSES[UNKNOWN_CODE])
    llm_label = CAUSES.get(llm_code, CAUSES[UNKNOWN_CODE])

    # Guardar resultados
    df.at[idx, OUTPUT_COLUMN_CODE] = code
    df.at[idx, OUTPUT_COLUMN_LABEL] = label
    df.at[idx, OUTPUT_COLUMN_CONF] = round(float(conf), 4)
    df.at[idx, OUTPUT_COLUMN_REVIEW] = bool(needs_review)
    df.at[idx, OUTPUT_COLUMN_STATUS] = status
    df.at[idx, OUTPUT_COLUMN_CODE_LLM] = llm_code
    df.at[idx, OUTPUT_COLUMN_LABEL_LLM] = llm_label
    df.at[idx, OUTPUT_COLUMN_CONF_LLM] = round(float(llm_conf), 4)
    df.at[idx, OUTPUT_COLUMN_RULE] = rule_applied

    print(
        f"[{n}/{len(df)}] "
        f"FINAL={code} | "
        f"LLM={llm_code} | "
        f"CONF={float(conf):.2f} | "
        f"RULE={rule_applied} | "
        f"STATUS={status}"
    )

    # Guardado parcial COMPLETO para no perder progreso
    if n % 10 == 0:
        df.to_excel("backup_clasificacion_parcial.xlsx", index=False)
        print(f"Guardado parcial completo: {n} filas procesadas")

    time.sleep(SLEEP_SECONDS)

# -----------------------------
# 3. Reforzar consistencia final
# -----------------------------

df[OUTPUT_COLUMN_CODE] = (
    df[OUTPUT_COLUMN_CODE]
    .apply(lambda x: label_to_code(x) or UNKNOWN_CODE)
)

df[OUTPUT_COLUMN_CODE] = df[OUTPUT_COLUMN_CODE].where(
    df[OUTPUT_COLUMN_CODE].isin(VALID_CODES),
    UNKNOWN_CODE
)

df[OUTPUT_COLUMN_LABEL] = df[OUTPUT_COLUMN_CODE].map(CAUSES)

df[OUTPUT_COLUMN_CODE_LLM] = (
    df[OUTPUT_COLUMN_CODE_LLM]
    .apply(lambda x: label_to_code(x) or UNKNOWN_CODE)
)

df[OUTPUT_COLUMN_CODE_LLM] = df[OUTPUT_COLUMN_CODE_LLM].where(
    df[OUTPUT_COLUMN_CODE_LLM].isin(VALID_CODES),
    UNKNOWN_CODE
)

df[OUTPUT_COLUMN_LABEL_LLM] = df[OUTPUT_COLUMN_CODE_LLM].map(CAUSES)

# -----------------------------
# 4. Crear final_cause_code y final_cause_for_analysis
# -----------------------------

df[FINAL_CAUSE_CODE_COLUMN] = df[OUTPUT_COLUMN_CODE]
df[FINAL_CAUSE_COLUMN] = df[OUTPUT_COLUMN_LABEL]

# -----------------------------
# 4.1 Verificar errores de API
# -----------------------------

api_errors = (df[OUTPUT_COLUMN_STATUS] == "API_ERROR").sum()
print("\nErrores de API:", api_errors)

if api_errors > 0:
    print("ADVERTENCIA: Hay filas con API_ERROR. No las interpretes como clasificación válida.")

# -----------------------------
# 5. Diagnóstico de consistencia
# -----------------------------

labels_that_are_codes = (
    df[OUTPUT_COLUMN_LABEL]
    .astype(str)
    .str.strip()
    .str.upper()
    .isin(VALID_CODES)
    .sum()
)

codes_outside_catalog = (
    ~df[OUTPUT_COLUMN_CODE].isin(VALID_CODES)
).sum()

print("\nDiagnóstico de consistencia:")
print("- predicted_label con letras A-H:", labels_that_are_codes)
print("- Code_cause fuera de A-H:", codes_outside_catalog)

if labels_that_are_codes > 0:
    raise ValueError("predicted_label todavía contiene letras. Revisar mapeo.")

if codes_outside_catalog > 0:
    raise ValueError("Code_cause contiene códigos fuera de A-H.")

# -----------------------------
# 6. Exportar SOLO columnas solicitadas
# -----------------------------

columnas_finales = [
    "repo",
    "issue_number",
    "comment_id",
    "comment_author",
    "comment_created_at",
    "comment_body_raw",
    "comment_body_clean_final",

    # Gold / manual
    "Code_cause",
    GOLD_COLUMN,

    # Predicción final después de reglas
    OUTPUT_COLUMN_CODE,
    OUTPUT_COLUMN_LABEL,
    "final_cause_code",
    "final_cause_for_analysis",

    # Predicción LLM original antes de reglas
    "Code_cause_llm",
    "predicted_label_llm",
    OUTPUT_COLUMN_CONF_LLM,

    # Confianza, estado y reglas
    "Confidence",
    OUTPUT_COLUMN_REVIEW,
    OUTPUT_COLUMN_STATUS,
    "Priority_rule_applied",
]

faltantes = [c for c in columnas_finales if c not in df.columns]

if faltantes:
    print("\nColumnas faltantes:")
    print(faltantes)

columnas_existentes = [c for c in columnas_finales if c in df.columns]

df_export = df[columnas_existentes].copy()

OUTPUT_FILE = "dataset_mapeo.xlsx"

df_export.to_excel(OUTPUT_FILE, index=False)

# ============================================================
# FORMATO EXCEL: AJUSTE DE COLUMNAS Y ALINEACIÓN
# ============================================================

from openpyxl import load_workbook
from openpyxl.styles import Alignment

wb = load_workbook(OUTPUT_FILE)
ws = wb.active

for column_cells in ws.columns:

    max_length = 0

    for cell in column_cells:

        cell.alignment = Alignment(
            wrap_text=True,
            vertical="top",
            horizontal="left"
        )

        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass

    adjusted_width = min(max(max_length + 5, 20), 80)

    ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = 60

wb.save(OUTPUT_FILE)

print("\nFormato Excel aplicado correctamente.")
print("\nProceso completado.")
print(f"Archivo generado: {OUTPUT_FILE}")

print("\nColumnas exportadas:")
print(df_export.columns.tolist())

print("\nCantidad de columnas exportadas:")
print(len(df_export.columns))

print("\nShape final:")
print(df_export.shape)

display(df_export.head(10))


# ---

# Paso 12
df[[
    "comment_body_clean_final",
    "Code_cause_llm",
    "predicted_label_llm",
    "final_cause_code",
    "final_cause_for_analysis",
    "Priority_rule_applied"
]].head(30)

# ---

# Paso 13
print((df["final_cause_code"] != df["Code_cause_llm"]).value_counts())
print(df["final_cause_for_analysis"].value_counts())

# ---

# Paso 14
override_cases = df[
    df["Priority_rule_applied"].astype(str).str.contains(
        "override",
        case=False,
        na=False
    )
]

print("Total casos con override/rechazo:", len(override_cases))

print("\nTipos de regla:")
print(override_cases["Priority_rule_applied"].value_counts())

override_cases[[
    "comment_body_clean_final",
    "Code_cause_llm",
    "predicted_label_llm",
    "final_cause_code",
    "final_cause_for_analysis",
    "Priority_rule_applied"
]].head(100)

# ---

# paso 15
# ============================================
# VALIDAR DISTRIBUCIÓN DE PREDICCIONES
# ============================================
print("Distribución por código predicho:")
print(df[OUTPUT_COLUMN_CODE].value_counts(dropna=False).sort_index())

print("\nDistribución por etiqueta predicha:")
print(df[OUTPUT_COLUMN_LABEL].value_counts(dropna=False))


# ---

# paso 16
# ============================================
# EVALUACIÓN CONTRA Causa_Normalizada
# ============================================
# Solo funciona si Causa_Normalizada contiene etiquetas gold manuales del catálogo de 8 causas.

RESULT_FILE = OUTPUT_FILE
df_eval = pd.read_excel(RESULT_FILE)

if GOLD_COLUMN not in df_eval.columns:
    print(f"No existe la columna gold '{GOLD_COLUMN}'. No se puede evaluar.")
else:
    df_eval["true_code"] = df_eval[GOLD_COLUMN].apply(label_to_code)
    df_eval[OUTPUT_COLUMN_CODE] = df_eval[OUTPUT_COLUMN_CODE].astype(str).str.strip().str.upper()
    df_eval_valid = df_eval.dropna(subset=["true_code", OUTPUT_COLUMN_CODE]).copy()
    df_eval_valid = df_eval_valid[df_eval_valid["true_code"].isin(VALID_CODES)]
    df_eval_valid = df_eval_valid[df_eval_valid[OUTPUT_COLUMN_CODE].isin(VALID_CODES)]

    print("Filas totales clasificadas:", len(df_eval))
    print("Filas evaluables:", len(df_eval_valid))
    print("Filas sin gold o no mapeables:", len(df_eval) - len(df_eval_valid))

    if len(df_eval_valid) > 0:
        acc = accuracy_score(df_eval_valid["true_code"], df_eval_valid[OUTPUT_COLUMN_CODE])
        print(f"\nAccuracy por código: {acc:.2%}")
        ordered_codes = sorted(VALID_CODES)
        print("\nReporte por código:")
        print(classification_report(
            df_eval_valid["true_code"],
            df_eval_valid[OUTPUT_COLUMN_CODE],
            labels=ordered_codes,
            target_names=[f"{c}: {CAUSES[c]}" for c in ordered_codes],
            zero_division=0
        ))
        cm = confusion_matrix(df_eval_valid["true_code"], df_eval_valid[OUTPUT_COLUMN_CODE], labels=ordered_codes)
        cm_df = pd.DataFrame(cm, index=[f"true_{c}" for c in ordered_codes], columns=[f"pred_{c}" for c in ordered_codes])
        print("\nMatriz de confusión:")
        display(cm_df)
        errors = df_eval_valid[df_eval_valid["true_code"] != df_eval_valid[OUTPUT_COLUMN_CODE]].copy()
        print("\nErrores:", len(errors))
        display_cols = [TEXT_COLUMN, GOLD_COLUMN, "true_code", OUTPUT_COLUMN_CODE, OUTPUT_COLUMN_LABEL, OUTPUT_COLUMN_CONF, OUTPUT_COLUMN_STATUS]
        display_cols = [c for c in display_cols if c in errors.columns]
        display(errors[display_cols].head(30))
    else:
        print("No hay filas evaluables. Si Causa_Normalizada está vacía, esto es esperado.")


# ---

# paso 17
# ============================================
# REVISIÓN GUIADA DE POSIBLES ERRORES
# ============================================
if "df_eval_valid" in globals() and len(df_eval_valid) > 0:
    disagreements = df_eval_valid[df_eval_valid["true_code"] != df_eval_valid[OUTPUT_COLUMN_CODE]].copy()

    def possible_reason(row):
        text = str(row.get(TEXT_COLUMN, "")).lower()
        pred = row.get(OUTPUT_COLUMN_CODE)

        if pred == "A" and any(x in text for x in ["unclear", "confusing", "what do you mean", "more specifics", "clarify", "ambiguous"]):
            return "La predicción A puede estar justificada por ambigüedad o falta de claridad."
        if pred == "B" and any(x in text for x in ["back and forth", "conflicting", "same page", "handoff", "remaining comments", "coordination"]):
            return "La predicción B puede estar justificada por desalineación de coordinación o flujo de trabajo."
        if pred == "C" and any(x in text for x in ["compatibility", "third-party", "api", "migration", "async", "database", "sqlite", "oracle", "implementation"]):
            return "La predicción C puede estar justificada por complejidad técnica, compatibilidad o restricciones del sistema."
        if pred == "D" and any(x in text for x in ["triage", "approval", "ticket", "needs tests", "patch", "merge", "rebase", "status"]):
            return "La predicción D puede estar justificada por restricción procedimental o de workflow."
        if pred == "E" and any(x in text for x in ["unfair", "exhausting", "civil", "code of conduct", "frustrat"]):
            return "La predicción E puede estar justificada por tensión interpersonal o percepción de trato injusto."
        if pred == "F" and any(x in text for x in ["docs", "documentation", "release notes", "versionchanged", "deprecation", "tutorial"]):
            return "La predicción F puede estar justificada por documentación o transferencia de conocimiento."
        if pred == "G" and any(x in text for x in ["review", "validate", "confirm", "can you", "permission", "access", "ci", "help move"]):
            return "La predicción G puede estar justificada por dependencia de validación, acceso, herramientas o recursos."
        if pred == "H":
            return "La predicción H indica que no se identificó causa sociotécnica clara. Revisar si el comentario solo confirma, agradece o informa."
        return "Revisar manualmente."

    disagreements["review_note"] = disagreements.apply(possible_reason, axis=1)
    cols = [TEXT_COLUMN, GOLD_COLUMN, "true_code", OUTPUT_COLUMN_CODE, OUTPUT_COLUMN_LABEL, OUTPUT_COLUMN_CONF, OUTPUT_COLUMN_CODE_LLM, OUTPUT_COLUMN_RULE, "review_note"]
    cols = [c for c in cols if c in disagreements.columns]
    print("Total de desacuerdos para revisar:", len(disagreements))
    display(disagreements[cols].head(50))
else:
    print("Primero ejecuta la celda de evaluación y asegúrate de tener gold manual.")


# ---

# paso 18
# ============================================
# ANÁLISIS DE ERRORES FRECUENTES
# ============================================
if "df_eval_valid" in globals() and len(df_eval_valid) > 0:
    errors = df_eval_valid[df_eval_valid["true_code"] != df_eval_valid[OUTPUT_COLUMN_CODE]].copy()
    print("Distribución de predicciones erradas:")
    print(errors[OUTPUT_COLUMN_CODE].value_counts(dropna=False))
    print("\nMatriz de confusión simple:")
    confusion = pd.crosstab(df_eval_valid["true_code"], df_eval_valid[OUTPUT_COLUMN_CODE], rownames=["Real"], colnames=["Predicho"])
    display(confusion)
    display_cols = [TEXT_COLUMN, GOLD_COLUMN, "true_code", OUTPUT_COLUMN_CODE, OUTPUT_COLUMN_LABEL, OUTPUT_COLUMN_CONF, OUTPUT_COLUMN_REVIEW, OUTPUT_COLUMN_STATUS]
    display_cols = [c for c in display_cols if c in errors.columns]
    print("\nPrimeros errores:")
    display(errors[display_cols].head(30))
else:
    print("Primero ejecuta la celda de evaluación y asegúrate de tener gold manual.")


# ---

# paso 19
# ======================================
# VALIDACIÓN DEL MAPEO FINAL CORREGIDA
# ======================================

required_cols = [
    TEXT_COLUMN,
    "Code_cause_llm",
    "predicted_label_llm",
    "Confidence_llm",
    "Priority_rule_applied",
    "final_cause_code",
    "final_cause_for_analysis"
]

missing = [c for c in required_cols if c not in df.columns]

if missing:
    print("Faltan columnas:", missing)
else:
    print("Todas las columnas requeridas existen.")

# Validar códigos permitidos
valid_codes = set(CAUSES.keys())

invalid_codes = df[
    ~df["final_cause_code"].isin(valid_codes)
]

print("Registros con código final inválido:", len(invalid_codes))

# Validar etiquetas finales
df["expected_final_label"] = df["final_cause_code"].map(CAUSES)

label_mismatch = df[
    df["final_cause_for_analysis"] != df["expected_final_label"]
]

print("Registros con etiqueta final inconsistente:", len(label_mismatch))

# Verificar cuántas reglas cambiaron el LLM
df["rule_changed_llm"] = df["final_cause_code"] != df["Code_cause_llm"]

print("\nCambios regla vs LLM:")
print(df["rule_changed_llm"].value_counts())

print("\nDistribución final:")
print(df["final_cause_for_analysis"].value_counts())

# Ver muestra de cambios
df[df["rule_changed_llm"]][[
    TEXT_COLUMN,
    "Code_cause_llm",
    "predicted_label_llm",
    "final_cause_code",
    "final_cause_for_analysis",
    "Priority_rule_applied"
]].head(20)

# ---


# Paso 20
# ============================================
# EXPORTAR FILAS PARA REVISIÓN MANUAL
# ============================================
REVIEW_FILE = "filas_para_revision_manual.xlsx"

if OUTPUT_COLUMN_REVIEW in df.columns:
    review_rows = df[df[OUTPUT_COLUMN_REVIEW].astype(bool)].copy()
    review_rows.to_excel(REVIEW_FILE, index=False)
    print(f"Filas para revisión manual: {len(review_rows)}")
    print(f"Archivo guardado: {REVIEW_FILE}")
    display_cols = [TEXT_COLUMN, OUTPUT_COLUMN_CODE, OUTPUT_COLUMN_LABEL, OUTPUT_COLUMN_CONF, OUTPUT_COLUMN_STATUS]
    display_cols = [c for c in display_cols if c in review_rows.columns]
    display(review_rows[display_cols].head(30))
else:
    print("No existe la columna de revisión manual. Ejecuta primero la clasificación.")


# ---

# Paso 21
# ============================================
# RESUMEN FINAL Y EXPORTS — VERSIÓN CORREGIDA
# NO sobrescribe final_cause_code/final_cause_for_analysis desde LLM
# ============================================

# ------------------------------------------------------------
# 1. Validar que existan columnas finales reales
# ------------------------------------------------------------

required_final_cols = [
    FINAL_CAUSE_CODE_COLUMN,
    FINAL_CAUSE_COLUMN,
    OUTPUT_COLUMN_CODE,
    OUTPUT_COLUMN_LABEL,
    OUTPUT_COLUMN_CODE_LLM,
    OUTPUT_COLUMN_LABEL_LLM,
    OUTPUT_COLUMN_RULE,
]

missing = [c for c in required_final_cols if c not in df.columns]
if missing:
    raise ValueError(
        "Faltan columnas necesarias. Ejecuta primero la celda de clasificación completa. "
        f"Columnas faltantes: {missing}"
    )

# ------------------------------------------------------------
# 2. Normalizar SIN cambiar la decisión final
# ------------------------------------------------------------

df[FINAL_CAUSE_CODE_COLUMN] = df[FINAL_CAUSE_CODE_COLUMN].apply(
    lambda x: label_to_code(x) or UNKNOWN_CODE
)

df[FINAL_CAUSE_CODE_COLUMN] = df[FINAL_CAUSE_CODE_COLUMN].where(
    df[FINAL_CAUSE_CODE_COLUMN].isin(VALID_CODES),
    UNKNOWN_CODE
)

df[FINAL_CAUSE_COLUMN] = df[FINAL_CAUSE_CODE_COLUMN].map(CAUSES)

df[OUTPUT_COLUMN_CODE] = df[OUTPUT_COLUMN_CODE].apply(
    lambda x: label_to_code(x) or UNKNOWN_CODE
)

df[OUTPUT_COLUMN_LABEL] = df[OUTPUT_COLUMN_CODE].map(CAUSES)

df[OUTPUT_COLUMN_CODE_LLM] = df[OUTPUT_COLUMN_CODE_LLM].apply(
    lambda x: label_to_code(x) or UNKNOWN_CODE
)

df[OUTPUT_COLUMN_LABEL_LLM] = df[OUTPUT_COLUMN_CODE_LLM].map(CAUSES)

# ------------------------------------------------------------
# 3. Diagnóstico clave: verificar si las reglas realmente cambiaron algo
# ------------------------------------------------------------

df["rule_changed_llm"] = df[FINAL_CAUSE_CODE_COLUMN] != df[OUTPUT_COLUMN_CODE_LLM]

print("===== DIAGNÓSTICO FINAL VS LLM =====")
print("Final distinto al LLM:")
print(df["rule_changed_llm"].value_counts(dropna=False))

print("\nReglas aplicadas:")
print(df[OUTPUT_COLUMN_RULE].value_counts(dropna=False).head(30))

print("\nCasos donde la regla cambió la salida del LLM:")
cols_changed = [
    TEXT_COLUMN,
    FINAL_CAUSE_CODE_COLUMN,
    FINAL_CAUSE_COLUMN,
    OUTPUT_COLUMN_CODE_LLM,
    OUTPUT_COLUMN_LABEL_LLM,
    OUTPUT_COLUMN_CONF,
    OUTPUT_COLUMN_RULE,
]
cols_changed = [c for c in cols_changed if c in df.columns]
display(df.loc[df["rule_changed_llm"], cols_changed].head(30))

# ------------------------------------------------------------
# 4. Normalizar confianza
# ------------------------------------------------------------

if OUTPUT_COLUMN_CONF in df.columns:
    df[OUTPUT_COLUMN_CONF] = (
        df[OUTPUT_COLUMN_CONF]
        .astype(str)
        .str.replace(",", ".", regex=False)
    )
    df[OUTPUT_COLUMN_CONF] = pd.to_numeric(df[OUTPUT_COLUMN_CONF], errors="coerce")
else:
    df[OUTPUT_COLUMN_CONF] = None

# ------------------------------------------------------------
# 5. Preparar gold solo para evaluación
# ------------------------------------------------------------

if GOLD_COLUMN in df.columns:
    df["true_code"] = df[GOLD_COLUMN].apply(label_to_code)
    df["evaluable"] = df["true_code"].isin(VALID_CODES)
else:
    df["true_code"] = None
    df["evaluable"] = False

# ------------------------------------------------------------
# 6. Calcular match contra gold, sin sobrescribir final
# ------------------------------------------------------------

df["match"] = None
mask_eval = df["evaluable"] == True

df.loc[mask_eval, "match"] = (
    df.loc[mask_eval, "true_code"] == df.loc[mask_eval, FINAL_CAUSE_CODE_COLUMN]
)

# ------------------------------------------------------------
# 7. Marcar revisión
# ------------------------------------------------------------

df[OUTPUT_COLUMN_REVIEW] = df[OUTPUT_COLUMN_CONF].fillna(0) < CONFIDENCE_THRESHOLD

df.loc[
    (df["evaluable"] == True) & (df["match"] == False),
    OUTPUT_COLUMN_REVIEW
] = True

df[OUTPUT_COLUMN_STATUS] = df[OUTPUT_COLUMN_REVIEW].map(
    {True: "REVIEW", False: "ACCEPTED"}
)

# No perder estados relevantes producidos por classify si ya existían.
# Si quieres conservar RULE_OVERRIDE/RULE_CONFIRMED, no vuelvas a ejecutar esta celda
# antes de exportar o crea una columna separada:
df["review_state"] = df[OUTPUT_COLUMN_STATUS]

# ------------------------------------------------------------
# 8. DataFrames de control
# ------------------------------------------------------------

df_evaluables = df[df["evaluable"] == True].copy()
df_sin_gold = df[df["evaluable"] == False].copy()
df_errors = df_evaluables[df_evaluables["match"] == False].copy()
df_doubtful = df[df[OUTPUT_COLUMN_CONF].fillna(0) < CONFIDENCE_THRESHOLD].copy()

resumen_por_causa = (
    df[FINAL_CAUSE_COLUMN]
    .value_counts(dropna=False)
    .rename_axis("causa_final")
    .reset_index(name="cantidad")
)

accuracy = df_evaluables["match"].mean() if len(df_evaluables) > 0 else None

print("\n===== RESUMEN =====")
print(f"Total procesados: {len(df)}")
print(f"Evaluables con causa real: {len(df_evaluables)}")
print(f"Sin causa real o sin gold válido: {len(df_sin_gold)}")
print(f"Errores evaluables: {len(df_errors)}")
print(f"Dudosos por confianza < {CONFIDENCE_THRESHOLD}: {len(df_doubtful)}")

if accuracy is not None:
    print(f"Accuracy general en evaluables: {accuracy:.4f}")
else:
    print("Accuracy general: no calculada porque no hay gold manual válido.")

print("\nDistribución final:")
display(resumen_por_causa)


# ---

# paso 22
# ============================================
# EXPORTACIÓN FINAL ÚNICA — DATASET_MAPEO — VERSIÓN CORREGIDA
# No recalcula final_cause_code desde LLM.
# Exporta trazabilidad completa: gold/manual, final con reglas, LLM y regla aplicada.
# ============================================

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FINAL_CLEAN_FILE = "dataset_mapeo.xlsx"
OUTPUT_FILE = FINAL_CLEAN_FILE

# ------------------------------------------------------------
# 1. Validar columnas finales ya calculadas por classify()
# ------------------------------------------------------------

required_export_cols = [
    FINAL_CAUSE_CODE_COLUMN,
    FINAL_CAUSE_COLUMN,
    OUTPUT_COLUMN_CODE,
    OUTPUT_COLUMN_LABEL,
    OUTPUT_COLUMN_CODE_LLM,
    OUTPUT_COLUMN_LABEL_LLM,
    OUTPUT_COLUMN_CONF,
    OUTPUT_COLUMN_RULE,
]

missing = [c for c in required_export_cols if c not in df.columns]
if missing:
    raise ValueError(
        "No exportes todavía. Faltan columnas de clasificación final. "
        f"Ejecuta primero la celda de clasificación. Faltan: {missing}"
    )

# Normalizar, pero NO reemplazar desde LLM.
df[FINAL_CAUSE_CODE_COLUMN] = df[FINAL_CAUSE_CODE_COLUMN].apply(
    lambda x: label_to_code(x) or UNKNOWN_CODE
)
df[FINAL_CAUSE_CODE_COLUMN] = df[FINAL_CAUSE_CODE_COLUMN].where(
    df[FINAL_CAUSE_CODE_COLUMN].isin(VALID_CODES),
    UNKNOWN_CODE
)
df[FINAL_CAUSE_COLUMN] = df[FINAL_CAUSE_CODE_COLUMN].map(CAUSES)

# Diagnóstico obligatorio antes de exportar
print("Diagnóstico antes de exportar:")
print("Final distinto al LLM:")
print((df[FINAL_CAUSE_CODE_COLUMN] != df[OUTPUT_COLUMN_CODE_LLM]).value_counts(dropna=False))

print("\nReglas aplicadas principales:")
print(df[OUTPUT_COLUMN_RULE].value_counts(dropna=False).head(20))

# ------------------------------------------------------------
# 2. Columnas finales para dataset_mapeo.xlsx
# ------------------------------------------------------------

columnas_finales = [
    "repo",
    "issue_number",
    "comment_id",
    "comment_author",
    "comment_created_at",
    "comment_body_raw",
    "comment_body_clean_final",

    # Gold/manual si existe
    "Code_cause",
    GOLD_COLUMN,
    "true_code",
    "evaluable",
    "match",

    # Predicción final después de reglas
    FINAL_CAUSE_CODE_COLUMN,
    FINAL_CAUSE_COLUMN,

    # Predicción final interna equivalente, generada en classify()
    OUTPUT_COLUMN_CODE,
    OUTPUT_COLUMN_LABEL,

    # LLM puro antes de reglas
    OUTPUT_COLUMN_CODE_LLM,
    OUTPUT_COLUMN_LABEL_LLM,
    OUTPUT_COLUMN_CONF_LLM,

    # Confianza, revisión y trazabilidad
    OUTPUT_COLUMN_CONF,
    OUTPUT_COLUMN_REVIEW,
    OUTPUT_COLUMN_STATUS,
    OUTPUT_COLUMN_RULE,
    "rule_changed_llm",
]

columnas_finales = list(dict.fromkeys(columnas_finales))
columnas_finales = [c for c in columnas_finales if c in df.columns]

df_final_limpio = df[columnas_finales].copy()

# ------------------------------------------------------------
# 3. Exportar Excel final
# ------------------------------------------------------------

df_final_limpio.to_excel(FINAL_CLEAN_FILE, index=False)

# ------------------------------------------------------------
# 4. Formato estable
# ------------------------------------------------------------

wb = load_workbook(FINAL_CLEAN_FILE)
ws = wb.active
ws.title = "dataset_mapeo"
ws.freeze_panes = "A2"

header_fill = PatternFill("solid", fgColor="D9EAF7")
header_font = Font(bold=True)

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

ws.auto_filter.ref = ws.dimensions

for col_idx, column_cells in enumerate(ws.columns, start=1):
    column_letter = get_column_letter(col_idx)
    header = str(column_cells[0].value or "")

    if header in ["comment_body_raw", "comment_body_clean_final"]:
        width = 70
    elif header in [FINAL_CAUSE_COLUMN, OUTPUT_COLUMN_LABEL, OUTPUT_COLUMN_LABEL_LLM, OUTPUT_COLUMN_RULE]:
        width = 45
    elif header in ["repo", "comment_author", "comment_created_at"]:
        width = 22
    else:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        width = min(max(max_length + 3, 12), 35)

    ws.column_dimensions[column_letter].width = width

ws.row_dimensions[1].height = 28
for row_idx in range(2, ws.max_row + 1):
    ws.row_dimensions[row_idx].height = 80

wb.save(FINAL_CLEAN_FILE)

print("Archivo final generado correctamente:")
print(FINAL_CLEAN_FILE)
print("\nColumnas exportadas:")
print(df_final_limpio.columns.tolist())
print("\nFilas y columnas:")
print(df_final_limpio.shape)

display(df_final_limpio.head(10))


# ---

# paso 23
test_comments = [
    "Can you review?",
    "What do you think?",
    "Any feedback?",
    "/workflow_ci",
    "I applied your suggestions.",
    "We cannot proceed until a maintainer validates this.",
    "Repository permissions are blocking contributors.",
]
for t in test_comments:
    result = classify(t)

    print("\nComentario:", t)
    print("FINAL:", result[0], "-", result[1])
    print("LLM:", result[5], "-", result[6])
    print("RULE:", result[8])
    print("-" * 80)

# ---

# paso 24
for c in test_comments:
    result = classify(c)

    print("Comentario:", c)
    print("Código final:", result[0])
    print("Causa final:", result[1])
    print("Confianza:", result[2])
    print("Regla aplicada:", result[8])
    print("-" * 80)

# ---

# paso 25
# ============================================
# VALIDACIÓN LOCAL DE REGLAS SIN API
# ============================================

# Ajustes necesarios antes de probar
if r"\btesting\b" not in VALIDATION_EVIDENCE_PATTERNS:
    VALIDATION_EVIDENCE_PATTERNS.append(r"\btesting\b")

BOT_TEMPLATE_PATTERNS = [
    r"automated_bot_message",
    r"workflow_ci",
    r"ci_status_reference",
    r"flaky tests guide",
    r"retesting failed pr",
    r"pull-request has been approved",
    r"lgtm label has been added",
    # Se elimina "this issue is currently awaiting triage"
    # porque puede representar triage/procedimiento => D
]

critical_tests = [
    (
        "This issue is currently awaiting triage. If a SIG or subproject determines this is relevant, they will accept it by applying the label.",
        "D"
    ),
    (
        "/workflow_ci",
        "H"
    ),
    (
        "Would we allow someone to add amd dra testing also? Are we hoping to support more expensive gpu resources like A100,H100 or MI250x with DRA?",
        "G"
    ),
    (
        "Using openTextDocument results in losing the tracking of problems after file close due to not replaced diagnostic entries.",
        "C"
    ),
    (
        "Can you review?",
        "H"
    ),
    (
        "I applied your suggestions.",
        "H"
    ),
]

for comment, expected in critical_tests:
    code, conf, rule = apply_priority_rules(comment, "H", 0.50)
    ok = "OK" if code == expected else "REVISAR"

    print("Esperado:", expected, "| Obtenido:", code, "|", ok)
    print("Regla:", rule)
    print(comment[:140])
    print("-" * 80)

# ---

# paso 26
# ============================================
# VALIDACIÓN LOCAL CRÍTICA SIN API
# ============================================

critical_tests = [
    {"comment": "This issue is currently awaiting triage.", "expected": "D"},
    {"comment": "/workflow_ci", "expected": "H"},
    {"comment": "Can you review?", "expected": "H"},
    {"comment": "We cannot proceed until maintainers validate this.", "expected": "G"},
    {"comment": "VSCode should clear diagnostics for documents that aren't open.", "expected": "C"},
]

for item in critical_tests:
    code, conf, rule = apply_priority_rules(
        item["comment"],
        "H",
        0.50
    )

    ok = "OK" if code == item["expected"] else "REVISAR"

    print("Comentario:", item["comment"])
    print("Esperado:", item["expected"])
    print("Obtenido:", code)
    print("Confianza:", conf)
    print("Regla:", rule)
    print("Resultado:", ok)
    print("-" * 80)

# ---


# paso 27
for t in critical_tests:

    result = classify(t["comment"])

    print("\n========================")
    print("Comentario:")
    print(t["comment"])

    print("\nEsperado:", t["expected"])
    print("Predicho:", result[0])

    print("Causa:", result[1])
    print("Regla:", result[8])

# ---

# paso 28
for c in [
    "/workflow_coordination",
    "This issue is currently awaiting triage.",
    "This PR may require API review.",
    "There is a race condition during stress testing.",
    "I count no less than 6 PRs for this. Who is comparing impl options?",
]:
    print(c, "=>", apply_priority_rules(c, "H", 0.5))

# ---

# paso 29
print(df["Priority_rule_applied"].value_counts(dropna=False).head(30))

print(df["final_cause_for_analysis"].value_counts(dropna=False))

# ---

# paso 30
mask_h = (
    df["final_cause_for_analysis"]
    .astype(str)
    .str.strip()
    .eq("No identifiable sociotechnical cause")
)

sample_h = df.loc[
    mask_h,
    [TEXT_COLUMN, "Confidence", "Priority_rule_applied"]
].sample(50, random_state=42)

for i, row in sample_h.iterrows():
    print("\n----------------------------")
    print(row[TEXT_COLUMN])
    print("CONF:", row["Confidence"])
    print("RULE:", row["Priority_rule_applied"])

# ---

# paso 31
mask_h = (
    df["final_cause_for_analysis"]
    .astype(str)
    .str.strip()
    .eq("No identifiable sociotechnical cause")
)

df_h_review = df.loc[
    mask_h,
    [
        TEXT_COLUMN,
        "Confidence",
        "Priority_rule_applied",
        "Code_cause_llm",
        "predicted_label_llm",
    ]
].copy()

print("Total H:", len(df_h_review))

df_h_review.sample(50, random_state=42)

# ---

# paso 32
df[
    [
        TEXT_COLUMN,
        "final_cause_code",
        "final_cause_for_analysis",
        "Code_cause_llm",
        "predicted_label_llm",
        "Priority_rule_applied",
    ]
].sample(20, random_state=42)

# ---

# Paso 33
print((df["final_cause_code"] == df["Code_cause_llm"]).value_counts())
print((df["final_cause_for_analysis"] == df["predicted_label_llm"]).value_counts())

# ---

# Paso 34
print(df.columns.tolist())

for col in df.columns:
    if "code" in col.lower() or "cause" in col.lower() or "label" in col.lower():
        print(col, "=>", df[col].dropna().astype(str).unique()[:10])

# ---

# Paso 35
import os

for f in os.listdir():
    if f.endswith((".xlsx", ".csv")):
        print(f)

# ---

# PASO 36
# ============================================
# VALIDAR ARCHIVO FINAL EXPORTADO
# ============================================

df_final = pd.read_excel("dataset_mapeo.xlsx")

print(df_final.columns.tolist())

for col in [
    "Code_cause_llm",
    "predicted_label_llm",
    "predicted_code",
    "predicted_label",
    "final_cause_code",
    "final_cause_for_analysis"
]:
    if col in df_final.columns:
        print("\n", col)
        print(df_final[col].dropna().astype(str).unique()[:10])