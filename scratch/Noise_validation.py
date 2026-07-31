import pandas as pd

df = pd.read_excel("/content/dataset_clean_final.xlsx")

print("Filas:", df.shape[0])
print("Columnas:", df.columns.tolist())

# Definir la columna de texto

TEXT_COLUMN = "comment_body_clean_final"

# Crear copia de trabajo

df_ruido = df.copy()

# ---

# ============================================================
# PASO 2. CREAR COLUMNA DE TEXTO NORMALIZADA
# ============================================================

df_ruido["texto"] = (
    df_ruido[TEXT_COLUMN]
    .fillna("")
    .astype(str)
    .str.strip()
)

# ============================================================
# PASO 3. INICIALIZAR CLASIFICACIÓN DE RUIDO
# Todo inicia como useful.
# Luego las reglas cambiarán a hard_noise u operational_noise.
# ============================================================

df_ruido["noise_level"] = "useful"

# ============================================================
# PASO 4. CREAR COLUMNA DE JUSTIFICACIÓN
# ============================================================

df_ruido["noise_reason"] = ""

# ============================================================
# PASO 5. VERIFICAR COLUMNAS CREADAS
# ============================================================

df_ruido[
    [
        TEXT_COLUMN,
        "texto",
        "noise_level",
        "noise_reason"
    ]
].head()

# ---

# ============================================================
# PASO 6. DETECTAR HARD_NOISE
# Ruido estructural, automático, placeholders, comandos puros,
# notificaciones automáticas o contenido sin densidad semántica
# ============================================================

import re

# ------------------------------------------------------------
# 6.1. Patrones exactos de hard_noise
# ------------------------------------------------------------

HARD_EXACT_PATTERNS = [
    # Metadata / placeholders estructurales
    r"^/?html_details_block\.?$",
    r"^/?html_image_reference\.?$",
    r"^/?html_meta_comment\.?$",
    r"^/?code_block_attached\.?$",
    r"^/?diff_attached\.?$",
    r"^/?error_log_attached\.?$",
    r"^/?log_reference\.?$",
    r"^/?hash_reference\.?$",
    r"^/?url_reference\.?$",
    r"^/?benchmark_reference\.?$",
    r"^/?ci_status_reference\.?$",
    r"^/?workflow_ci\.?$",
    r"^/?workflow_review\.?$",
    r"^/?workflow_coordination\.?$",
    r"^/?approval_notification\.?$",
    r"^/?automated_cla_check\.?$",
    r"^/?automated_bot_message\.?$",
    r"^/?inline_code\.?$",

    # Placeholders de usuarios sin contenido adicional
    r"^a github user is mentioned\.?$",
    r"^multiple github users are mentioned\.?$",
    r"^a github user is asked to review or check this\.?$",
    r"^multiple github users are asked to review or check this\.?$",

    # Comandos puros
    r"^/[a-zA-Z0-9_-]+$",
    r"^/[a-zA-Z0-9_-]+\s+[a-zA-Z0-9_.-]+$",

    # Hash puro
    r"^[a-f0-9]{7,40}$",

    # Comentarios vacíos o símbolos sin contenido
    r"^\s*$",
    r"^\?+$",
    r"^!+$",
    r"^\)+$",
    r"^\(+$",
    r"^eyes:?$",
    r"^also$",
    r"^node$",
    r"^\d+$",
]


# ------------------------------------------------------------
# 6.2. Patrones de frases automáticas o estructurales
# ------------------------------------------------------------

HARD_PHRASE_PATTERNS = [
    # Estados automáticos por hash
    r"^landed in\s+/?hash_reference\.?$",
    r"^fixed in\s+/?hash_reference\.?$",
    r"^merged in\s+/?hash_reference\.?$",
    r"^closed by\s+/?hash_reference\.?$",

    # CI / test bots
    r"the following test[s]? \*\*?failed\*\*?",
    r"full pr test history",
    r"your pr dashboard",
    r"rerun all failed tests",
    r"rerun all mandatory failed tests",
    r"please help us cut down on flakes",
    r"unknown cla label state",
    r"codecov .* report",

    # Notificaciones automáticas de labels / aprobación
    r"^lgtm label has been added\.?\s*/?html_details_block\.?$",
    r"^approved label has been added\.?\s*/?html_details_block\.?$",
    r"^pull-request has been approved\.?\s*/?html_details_block\.?$",
    r"^pull request has been approved\.?\s*/?html_details_block\.?$",
    r"^[a-z0-9_-]+ label has been added\.?\s*/?html_details_block\.?$",
    r"^[a-z0-9_-]+ label has been removed\.?\s*/?html_details_block\.?$",
    r"new changes are detected\.?\s*lgtm label has been (added|removed)",
    r"new changes are detected\.?\s*lgtm label has been",

    # Labels / permisos / comandos automáticos
    r"label\(s\).*cannot be applied",
    r"the label .* cannot be applied",
    r"repository doesn't have them",
    r"only github organization members can add the label",
    r"you can only set the release note label",
    r"release-note-label-needed",
    r"no release-note block was detected",
    r"github didn't allow me to request pr reviews",
    r"failed to re-open pr",
    r"state cannot be changed",
    r"those labels are not set on the issue",
    r"there are no sig labels on this issue",
    r"adding label .* merge commits",
    r"release-note-edit must be used with a release note block",

    # Plantillas automáticas de contributor/help
    r"this request has been marked as needing help from a contributor",
    r"this request has been marked as suitable for new contributors",
    r"guidelines\s+please ensure that the issue body includes",
    r"instructions for interacting with me using pr comments",

    # Plantillas automáticas de verificación / testing de PR
    r"verify that this patch is reasonable to test",
    r"i will not automatically test new commits in this pr",

    # Imágenes, memes o comandos visuales sin contenido analítico
    r"^a github user is mentioned\.?\s*!cat image\s*/?url_reference\s*/?html_details_block\.?$",
    r"^multiple github users are mentioned\.?\s*!cat image\s*/?url_reference\s*/?html_details_block\.?$",
    r"!cat image\s*/?url_reference",

    # Plantillas automáticas de triage
    r"this issue is currently awaiting triage",
    r"if a sig or subproject determines this is a relevant issue",
]


# ------------------------------------------------------------
# 6.3. Tokens de bajo valor para limpieza residual
# ------------------------------------------------------------

PLACEHOLDERS_TO_REMOVE = [
    "/html_details_block",
    "/html_image_reference",
    "/html_meta_comment",
    "/code_block_attached",
    "/diff_attached",
    "/error_log_attached",
    "/log_reference",
    "/hash_reference",
    "/url_reference",
    "/benchmark_reference",
    "/ci_status_reference",
    "/automated_cla_check",
    "/automated_bot_message",
    "/approval_notification",
    "/workflow_ci",
    "/workflow_review",
    "/workflow_coordination",
    "/inline_code",
    "/emoji_celebra",
    "/emoji_other",
]


GENERIC_USER_PLACEHOLDERS = [
    r"a github user is mentioned\.?",
    r"multiple github users are mentioned\.?",
    r"a github user is asked to review or check this\.?",
    r"multiple github users are asked to review or check this\.?",
]


LOW_VALUE_VISUAL_PATTERNS = [
    r"!cat image",
    r"!dog image",
    r"!gif",
    r"!meme",
]


# ------------------------------------------------------------
# 6.4. Marcadores automáticos al inicio del comentario
# ------------------------------------------------------------

AUTOMATED_PREFIX_MARKERS = [
    "/automated_bot_message",
    "automated_bot_message",
    "/approval_notification",
    "approval_notification",
    "/automated_cla_check",
    "automated_cla_check",
    "/workflow_ci",
    "workflow_ci",
    "/workflow_review",
    "workflow_review",
    "/workflow_coordination",
    "workflow_coordination",
    "/ci_status_reference",
    "ci_status_reference",
]


# ------------------------------------------------------------
# 6.5. Funciones auxiliares
# ------------------------------------------------------------

def normalize_text(texto):
    return str(texto).strip().lower()


def matches_any_fullmatch(texto, patterns):
    return any(
        re.fullmatch(pattern, texto, flags=re.DOTALL)
        for pattern in patterns
    )


def matches_any_search(texto, patterns):
    return any(
        re.search(pattern, texto, flags=re.DOTALL)
        for pattern in patterns
    )


def starts_with_automated_marker(texto):
    texto = texto.strip().lower()

    return any(
        texto.startswith(marker)
        for marker in AUTOMATED_PREFIX_MARKERS
    )


def is_only_slash_commands(texto):
    lineas = [line.strip() for line in texto.splitlines() if line.strip()]

    if not lineas:
        return False

    return all(
        re.fullmatch(r"/[a-zA-Z0-9_-]+(\s+[a-zA-Z0-9_.-]+)?", line)
        for line in lineas
    )


def remove_low_value_tokens(texto):
    texto_limpio = texto

    for placeholder in PLACEHOLDERS_TO_REMOVE:
        texto_limpio = texto_limpio.replace(placeholder, " ")

    for pattern in GENERIC_USER_PLACEHOLDERS:
        texto_limpio = re.sub(pattern, " ", texto_limpio)

    for pattern in LOW_VALUE_VISUAL_PATTERNS:
        texto_limpio = re.sub(pattern, " ", texto_limpio)

    texto_limpio = re.sub(r"\s+", " ", texto_limpio).strip()

    return texto_limpio


# ------------------------------------------------------------
# 6.6. Función principal hard_noise
# ------------------------------------------------------------

def es_hard_noise(texto):
    texto = normalize_text(texto)

    # 0. Mensajes automáticos que empiezan con marcador estructural
    if starts_with_automated_marker(texto):
        return True

    # 1. Placeholders, comandos, hashes o símbolos puros
    if matches_any_fullmatch(texto, HARD_EXACT_PATTERNS):
        return True

    # 2. Comentarios formados solo por comandos slash
    if is_only_slash_commands(texto):
        return True

    # 3. Plantillas automáticas, bots, labels, permisos o CI
    if matches_any_search(texto, HARD_PHRASE_PATTERNS):
        return True

    # 4. Comentarios que al quitar placeholders quedan vacíos o casi vacíos
    texto_limpio = remove_low_value_tokens(texto)

    if len(texto_limpio) <= 3:
        return True

    return False

# ---

# ============================================================
# APLICAR HARD_NOISE
# ============================================================

df_ruido["is_hard_noise"] = df_ruido["texto"].apply(es_hard_noise)

df_ruido.loc[
    df_ruido["is_hard_noise"],
    "noise_level"
] = "hard_noise"

df_ruido.loc[
    df_ruido["is_hard_noise"],
    "noise_reason"
] = (
    "Ruido estructural, automático, placeholder o comentario sin contenido interpretable"
)

# Resumen de clasificación después de hard_noise
print(df_ruido["noise_level"].value_counts())

# Revisar primeros ejemplos clasificados como hard_noise
display(
    df_ruido[df_ruido["noise_level"] == "hard_noise"][
        ["texto", "noise_level", "noise_reason"]
    ].head(30)
)

# Revisar muestra aleatoria de hard_noise
display(
    df_ruido[df_ruido["noise_level"] == "hard_noise"][
        ["texto", "noise_level", "noise_reason"]
    ].sample(
        min(30, len(df_ruido[df_ruido["noise_level"] == "hard_noise"])),
        random_state=42
    )
)

# ---

# ============================================================
# PASO 7. DETECTAR OPERATIONAL_NOISE
# Interacciones humanas mínimas de workflow, revisión o mantenimiento
# sin suficiente valor sociotécnico causal
# ============================================================

# ------------------------------------------------------------
# 7.1. Patrones exactos de operational_noise
# ------------------------------------------------------------

OPERATIONAL_EXACT_PATTERNS = [
    # Respuestas mínimas
    r"^thanks\.?$",
    r"^thanks!$",
    r"^thank you\.?$",
    r"^ok\.?$",
    r"^okay\.?$",
    r"^yes\.?$",
    r"^yeah\.?$",
    r"^no worries\.?$",
    r"^sounds good\.?$",
    r"^looks good\.?$",
    r"^all good!?$",
    r"^acknowledged\.?$",

    # Estados mínimos de cambio
    r"^done\.?$",
    r"^updated\.?$",
    r"^fixed\.?$",
    r"^resolved\.?$",
    r"^addressed\.?$",
    r"^squashed\.?$",
    r"^rebased\.?$",
    r"^rebase\.?$",
    r"^rebase done\.?$",
    r"^force-pushed\.?$",
    r"^amended\.?$",

    # Revisión mínima
    r"^ptal\.?$",
    r"^lgtm\.?$",
    r"^approved\.?$",
    r"^approve, thanks!?$",
    r"^looks good to me\.?$",
    r"^this one is looking good\.?$",
    r"^ready for review\.?$",
    r"^ready for another pass\.?$",
    r"^i am reviewing this pr\.?$",
    r"^reviewing new commits\.?$",

    # CI / estado operacional humano mínimo
    r"^ci is green.*$",
    r"^ci looks good.*$",
    r"^test was successful!?$",
    r"^tests? passed\.?$",
    r"^passes locally\.?$",
    r"^passed locally\.?$",
    r"^unrelated failure\.?$",
    r"^seems unrelated\.?$",
    r"^not a flaking test\.?$",
]


# ------------------------------------------------------------
# 7.2. Patrones de frases operacionales mínimas
# ------------------------------------------------------------

OPERATIONAL_PHRASE_PATTERNS = [
    # Rebase / conflictos / mantenimiento PR
    r"^needs rebase\.?$",
    r"^please rebase\.?$",
    r"^rebased to fix conflict.*$",
    r"^rebased and removed merge commit.*$",
    r"^i have rebased.*$",
    r"^resolved trivial conflict.*$",
    r"^resolved conflict.*$",
    r"^only solved conflicts.*$",
    r"^merge conflict.*$",

    # Squash / commits / ajustes mínimos
    r"^done,? thanks.*$",
    r"^fixed now\.?$",
    r"^should be fixed now\.?$",
    r"^i have revised it\.?$",
    r"^i revised it\.?$",
    r"^i updated it\.?$",
    r"^i fixed it\.?$",
    r"^i pushed it\.?$",
    r"^squashed.*$",
    r"^amended the commit.*$",
    r"^force-pushed.*$",

    # Comentarios/revisión ya atendidos
    r"^comments addressed\.?$",
    r"^all comments have been addressed\.?$",
    r"^adjusted all comments.*$",
    r"^i think i addressed all the comments\.?$",
    r"^i have addressed all the comments\.?$",
    r"^i believe i have resolved all review comments\.?$",
    r"^i believe your comments have been addressed\.?$",

    # Pings / solicitudes simples de revisión
    r"^friendly ping.*$",
    r"^gentle ping.*$",
    r"^ping for review.*$",
    r"^pinging .* for review.*$",
    r"^please take a look\.?$",
    r"^can you take a look\.?$",
    r"^can you please take a look\.?$",
    r"^could you please take a look\.?$",
    r"^could you help review\.?$",
    r"^please review\.?$",
    r"^review,? thanks!?$",
    r"^ptal\.?$",
    r"^ptal,? thanks!?$",
    r"^any update\??$",
    r"^any further feedback\??$",
    r"^does anything else still need to be done\??$",
    r"^is this still active\??$",

    # Asignación o toma de tarea mínima
    r"^can i work on this issue\??$",
    r"^can i fix\??$",
    r"^i want to work on this issue\.?$",
    r"^i would like to work on this\.?$",
    r"^i would like to work on this issue\.?$",
    r"^i'd like to work on this\.?$",
    r"^i'd like to work on this issue\.?$",
    r"^may i take this\??$",
    r"^may i take this issue\??$",
    r"^i will take this\.?$",
    r"^i will take this issue\.?$",
    r"^i will take up this issue\.?$",
    r"^i'll take this\.?$",
    r"^i'll take this issue\.?$",
    r"^i picked this up\.?$",
    r"^i picked up this issue\.?$",
    r"^i'm taking this\.?$",
    r"^i'm taking this issue\.?$",
    r"^i will take a look\.?$",
    r"^i'll take a look\.?$",
    r"^i'll give it a try\.?$",
    r"^i'll try\.?$",
    r"^i will try\.?$",

    # Compromiso breve de acción futura sin explicación técnica
    r"^(ok|okay|sure|alright|fine),?\s+i (will|can) do it( tomorrow| later)?\.?$",
    r"^i (will|can) do it( tomorrow| later)?\.?$",
    r"^i'll do it( tomorrow| later)?\.?$",

    # Respuesta breve después de una cita, sin explicación técnica adicional
    r'^based on the quotation:\s*".{1,150}"\s*(ok|okay|sure|alright|fine),?\s+i (will|can) do it( tomorrow| later)?\.?$',
    r'^based on the quotation:\s*".{1,150}"\s*i (will|can) do it( tomorrow| later)?\.?$',
    r'^based on the quotation:\s*".{1,150}"\s*i\'ll do it( tomorrow| later)?\.?$',

    # Espera / seguimiento mínimo
    r"^waiting for this.*$",
    r"^waiting for .* to be merged.*$",
    r"^this is ready for review\.?$",
    r"^ready to merge\.?$",
    r"^ready for merge\.?$",
    r"^ready to go\.?$",
    r"^this is good to go\.?$",
    r"^with .* merged, i think this is ready\.?$",
    r"^i think this is ready\.?$",

    # Cierre / reapertura / estado mínimo humano
    r"^closed\.$",
    r"^closed this pr\.?$",
    r"^closing this issue\.?$",
    r"^closed in favor of.*$",
    r"^reopening\.?$",
    r"^i'll close this pr\.?$",
    r"^i will close this pr\.?$",

    # Respuestas cortas con bajo valor causal
    r"^thanks for the update\.?$",
    r"^thanks for the information\.?$",
    r"^thanks for the review\.?$",
    r"^thanks for the feedback\.?$",
    r"^thanks for checking\.?$",
    r"^thanks for the clarification\.?$",
    r"^thank you for your guidance\.?$",
    r"^thank you for the review\.?$",
    r"^thank you for your feedback\.?$",
    r"^thank you for your comments\.?$",
    r"^thanks all\.?$",
    r"^thx all\.?$",

    # CI humano ligero / rerun manual / estado de pruebas
    r"^manually run the test job again\.?$",
    r"^re-running ci\.?$",
    r"^rerunning ci\.?$",
    r"^ci errors look legit\.?$",
    r"^(the )?ci is green now!?$",
    r"^test link:.*$",
    r"^passed in \d+.*$",
    r"^passed again.*$",

    # Solicitudes humanas breves para reactivar CI / pipeline / checks
    r"^can the pipeline be triggered again\??.*$",
    r"^can the pipeline be re-triggered\??.*$",
    r"^could the pipeline be triggered again\??.*$",
    r"^could the pipeline be re-triggered\??.*$",
    r"^please trigger the pipeline again\.?.*$",
    r"^please re-trigger the pipeline\.?.*$",
    r"^please rerun the pipeline\.?.*$",
    r"^please rerun ci\.?.*$",
    r"^please rerun the ci\.?.*$",
    r"^can ci be triggered again\??.*$",
    r"^could ci be triggered again\??.*$",
    r"^can the checks be triggered again\??.*$",
    r"^could the checks be triggered again\??.*$",
    r"^please rerun the checks\.?.*$",
    r"^checks are going now\.?$",
]


# ------------------------------------------------------------
# 7.3. Función principal operational_noise
# ------------------------------------------------------------

def es_operational_noise(texto):
    texto = normalize_text(texto)

    # Jerarquía: si es hard_noise, no puede ser operational_noise
    if es_hard_noise(texto):
        return False

    # Coincidencias exactas
    if matches_any_fullmatch(texto, OPERATIONAL_EXACT_PATTERNS):
        return True

    # Frases operacionales mínimas
    if matches_any_fullmatch(texto, OPERATIONAL_PHRASE_PATTERNS):
        return True

    return False

# ---

# ============================================================
# APLICAR OPERATIONAL_NOISE
# ============================================================

df_ruido["is_operational_noise"] = df_ruido["texto"].apply(es_operational_noise)

mask_operational = (
    df_ruido["is_operational_noise"] &
    (df_ruido["noise_level"] == "useful")
)

df_ruido.loc[
    mask_operational,
    "noise_level"
] = "operational_noise"

df_ruido.loc[
    mask_operational,
    "noise_reason"
] = (
    "Ruido operacional: interacción humana mínima de PR/revisión "
    "sin suficiente valor sociotécnico causal"
)

df_ruido["noise_level"].value_counts()

# ---

# ============================================================
# REVISAR EJEMPLOS DE OPERATIONAL_NOISE
# ============================================================

display(
    df_ruido[df_ruido["noise_level"] == "operational_noise"][
        ["texto", "noise_level", "noise_reason"]
    ].head(50)
)

# Muestra aleatoria para revisar variedad de casos
display(
    df_ruido[df_ruido["noise_level"] == "operational_noise"][
        ["texto", "noise_level", "noise_reason"]
    ].sample(
        min(50, len(df_ruido[df_ruido["noise_level"] == "operational_noise"])),
        random_state=42
    )
)

# ---

# ============================================================
# APLICAR OPERATIONAL_NOISE
# ============================================================

df_ruido["is_operational_noise"] = df_ruido["texto"].apply(es_operational_noise)

operational_mask = (
    (df_ruido["is_operational_noise"] == True) &
    (df_ruido["noise_level"] == "useful")
)

df_ruido.loc[
    operational_mask,
    "noise_level"
] = "operational_noise"

df_ruido.loc[
    operational_mask,
    "noise_reason"
] = (
    "Ruido operacional: interacción humana mínima de PR/revisión "
    "sin suficiente valor sociotécnico causal"
)

df_ruido["noise_level"].value_counts()

# ---

# Ver conteo

df_ruido["noise_level"].value_counts()



# ---

# ============================================================
# DEFINIR RAZÓN PARA USEFUL FINAL
# ============================================================

df_ruido.loc[
    df_ruido["noise_level"] == "useful",
    "noise_reason"
] = (
    "Contenido útil: comentario con valor técnico, comunicacional, "
    "organizacional o sociotécnico para análisis causal"
)

# ---

df_ruido["noise_level"].value_counts()

# ---

# Crear dataset limpio

df_limpio = df_ruido[
    df_ruido["noise_level"] == "useful"
].copy()

# ---

# ==========================================
# SELECCIONAR COLUMNAS FINALES
# ==========================================

columnas_finales = [
    "repo",
    "issue_number",
    "comment_id",
    "comment_author",
    "comment_created_at",
    "comment_body_raw",
    "comment_body_clean_final",

]

df_limpio = df_limpio[columnas_finales].copy()

# ---

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# EXPORTAR DATASET LIMPIO CON FORMATO
# ==========================================

output_file = "/content/dataset_limpio_final.xlsx"

# Exportar
df_limpio.to_excel(output_file, index=False)

# Abrir archivo
wb = load_workbook(output_file)
ws = wb.active
ws.title = "Dataset_Limpio"

# ----------------------------
# Estilo encabezado
# ----------------------------
header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")  # Azul oscuro
header_font = Font(color="FFFFFF", bold=True)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Aplicar formato al encabezado
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = border

# ----------------------------
# Formato de las celdas
# ----------------------------
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# ----------------------------
# Ajustar ancho automáticamente
# ----------------------------
for column_cells in ws.columns:
    length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
    ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 3, 50)

# ----------------------------
# Congelar encabezado
# ----------------------------
ws.freeze_panes = "A2"

# ----------------------------
# Agregar filtros
# ----------------------------
ws.auto_filter.ref = ws.dimensions

# Guardar
wb.save(output_file)

print("=" * 60)
print("DATASET FINAL LIBRE DE RUIDO")
print("=" * 60)
print(f"Comentarios útiles : {len(df_limpio):,}")
print(f"Columnas           : {len(df_limpio.columns)}")
print(f"Archivo generado   : {output_file}")
print("=" * 60)

# ---

print(df_limpio.shape)

# ---

# Verificar cantidad
df_noise["noise_level"].value_counts()


# ---

# ============================================================
# CONTROL DE RESIDUOS EN USEFUL
# ============================================================

residual_keywords = [
    "html_details_block",
    "html_image_reference",
    "html_meta_comment",
    "automated_bot_message",
    "automated_cla_check",
    "approval_notification",
    "codecov",
    "label has been added",
    "label has been removed",
    "pull-request has been approved",
    "verify that this patch is reasonable to test",
    "i will not automatically test new commits in this pr",
    "cat image",
    "please rerun",
    "pipeline be triggered",
    "checks are going now",
    "test link:",
    "ci:",
    "ci is green",
    "passed in",
    "lgtm",
    "ptal",
    "rebased",
    "force-pushed",
    "amended the commit",
]

pattern_residual = "|".join(residual_keywords)

df_residuos_useful = df_ruido[
    (df_ruido["noise_level"] == "useful") &
    (df_ruido["texto"].str.contains(pattern_residual, case=False, na=False))
]

print("Posibles residuos dentro de useful:", len(df_residuos_useful))

display(
    df_residuos_useful[
        ["texto", "noise_level", "noise_reason"]
    ].head(80)
)

# ---

df_ruido["noise_level"].value_counts()