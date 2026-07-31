# Cargar librerías y dataset
import os
import re
from pathlib import Path

import pandas as pd

# Buscar el dataset en rutas comunes sin depender de un solo entorno
candidate_names = [
    "dataset_master_clean2.csv",
    "dataset_limpio_inline_only.xlsx",
    "dataset_code_encapsulated.xlsx",
]

search_dirs = [
    Path.cwd(),
    Path("/mnt/data"),
    Path("/content"),
]

possible_paths = []
env_dataset = os.environ.get("DATASET_PATH")
if env_dataset:
    possible_paths.append(Path(env_dataset))

for base_dir in search_dirs:
    for name in candidate_names:
        possible_paths.append(base_dir / name)

# Quitar duplicados conservando el orden
seen = set()
ordered_paths = []
for p in possible_paths:
    p_str = str(p)
    if p_str not in seen:
        seen.add(p_str)
        ordered_paths.append(p)

file_path = next((p for p in ordered_paths if p.exists()), None)
if file_path is None:
    checked = "\n".join(f"- {p}" for p in ordered_paths)
    raise FileNotFoundError(
        "No se encontró un archivo de dataset compatible. Rutas revisadas:\n" + checked
    )

# Leer según la extensión del archivo
if file_path.suffix.lower() == ".csv":
    df = pd.read_csv(
    file_path,
    sep=None,
    engine="python",
    encoding="utf-8",
    on_bad_lines="skip"
)

elif file_path.suffix.lower() in [".xlsx", ".xls"]:
    df = pd.read_excel(file_path, engine="openpyxl")

else:
    raise ValueError(
        f"Formato no soportado: {file_path.suffix}"
    )

print(f"Dataset cargado desde: {file_path}")
print(df.shape)

# ---

# Mirar la columna de texto
df["comment_body_raw"].head(10)

# ---

# Define la regla para inline code
INLINE_RE = re.compile(r'`([^`\n]+)`')

def clean_inline_code(text):
    if not isinstance(text, str):
        return ""
    return INLINE_RE.sub("/inline_code", text)

print("Regla de inline code lista")

# Inline code sin backticks
INLINE_CODE_NO_BACKTICKS_RE = re.compile(
    r'(?m)^\s*["\']?\s*(?:'
    r'(?:private|public|protected|readonly|const|let|var|def|class|return|import|from)\b'
    r'|[A-Za-z_][A-Za-z0-9_]*\s*[:=]\s*[^,\n]+'
    r'|.*//.*'
    r')'
    r'.*$'
)

def clean_inline_code_without_backticks(text):
    if not isinstance(text, str):
        return ""
    return INLINE_CODE_NO_BACKTICKS_RE.sub("/inline_code", text)

# ---

# prueba básica de inline code
sample = df["comment_body_raw"].head(10).copy()

for i, text in enumerate(sample, 1):
    print("=" * 100)
    print(f"ROW {i}")
    print("ORIGINAL:")
    print(text)
    print("\nINLINE CLEANED:")
    print(clean_inline_code(text))
    print()


# ---

# Define etiquetas para los bloques y para menciones

CODE_BLOCK_TAG = "/Code_block_attached."
DIFF_BLOCK_TAG = "/Diff_attached."
ERROR_BLOCK_TAG = "/Error_log_attached."

USER_MENTION_SINGLE = "A GitHub user is mentioned."
USER_MENTION_MULTI = "Multiple GitHub users are mentioned."
USER_REVIEW_SINGLE = "A GitHub user is asked to review or check this."
USER_REVIEW_MULTI = "Multiple GitHub users are asked to review or check this."

# Keep these as internal categories if you want to track the mention type,
# but do not use them as full replacements for comments with extra content.
USER_LEADING_SINGLE = "A GitHub user is mentioned."
USER_LEADING_MULTI = "Multiple GitHub users are mentioned."

TEAM_MENTION_SINGLE = "A team from a GitHub organization is mentioned."
TEAM_MENTION_MULTI = "Multiple teams from GitHub organizations are mentioned."
TEAM_REVIEW_SINGLE = "A team from a GitHub organization is asked to review or check the comment."
TEAM_REVIEW_MULTI = "Multiple teams from GitHub organizations are asked to review or check the comment."

# Keep these as internal categories if you want to track the mention type,
# but do not use them as full replacements for comments with extra content.
TEAM_LEADING_SINGLE = "A team from a GitHub organization is mentioned."
TEAM_LEADING_MULTI = "Multiple teams from GitHub organizations are mentioned."

print("Output labels ready")

BOT_TAG = "/automated_bot_message"
print("Bot tag ready")

# ---


# Detectar bloques entre triple backticks, patrones de error y menciones
FENCED_BLOCK_RE = re.compile(
    r'```([A-Za-z0-9_+\-]*)\s*\n(.*?)\n\s*```',
    re.DOTALL
)

print("Fenced block pattern ready")

ERROR_SIGNALS = re.compile(
    r"""
    (?im)
    ^(?:ERROR|FAIL|FAILED|FATAL)[:\s]                # log lines starting with ERROR/FAIL/FAILED/FATAL
    |Traceback \(most recent call last\)            # Python traceback
    |={6,}\s*\n(?:FAIL|ERROR):                      # test separators followed by FAIL/ERROR
    |FAILED \(errors=\d+                            # unittest summary
    |\b\d+\s+(?:test|tests)\s+failed\b           # summary like "3 tests failed"
    |panic:\s+runtime\s+error                       # Go panic
    |^\s*[A-Za-z_][A-Za-z0-9_]*Error:                # ValueError:, TypeError:, AssertionError:
    |^\s*Exception:                                  # generic exception line
    |^\s*Caused by:                                  # Java style stack trace
    |\bexit code\s+\d+\b                          # exit code 1
    """,
    re.VERBOSE
)

ERROR_LOG_STRONG_RE = ERROR_SIGNALS

PROGRAMMING_LANGS = {
    "python", "py", "javascript", "js", "typescript", "ts",
    "java", "c", "cpp", "c++", "csharp", "cs", "go", "ruby", "rb",
    "php", "swift", "kotlin", "scala", "rust", "r", "sql",
    "bash", "shell", "sh", "zsh", "powershell", "ps1",
    "html", "css", "scss", "sass", "xml", "json", "yaml", "yml",
    "dockerfile", "makefile", "gradle"
}

DIFF_GIT_RE = re.compile(r'(?m)^\s*diff --git\b')
DIFF_HUNK_RE = re.compile(r'(?m)^\s*@@\s*-\d+(?:,\d+)?\s+\+\d+(?:,\d+)?\s*@@')
DIFF_FILE_OLD_RE = re.compile(r'(?m)^\s*---\s+[ab]/.+')
DIFF_FILE_NEW_RE = re.compile(r'(?m)^\s*\+\+\+\s+[ab]/.+')
DIFF_INDEX_RE = re.compile(r'(?m)^\s*index\s+[0-9a-f]+\.\.[0-9a-f]+\s+\d+')

print("Code classification signals ready")

def classify_fenced_block(lang, body):
    lang = (lang or "").strip().lower()

    if lang == "diff":
        return DIFF_BLOCK_TAG

    diff_signals = 0
    if DIFF_GIT_RE.search(body):
        diff_signals += 1
    if DIFF_HUNK_RE.search(body):
        diff_signals += 1
    if DIFF_FILE_OLD_RE.search(body):
        diff_signals += 1
    if DIFF_FILE_NEW_RE.search(body):
        diff_signals += 1
    if DIFF_INDEX_RE.search(body):
        diff_signals += 1

    if diff_signals >= 2:
        return DIFF_BLOCK_TAG

    if lang and lang in PROGRAMMING_LANGS:
        return CODE_BLOCK_TAG

    if ERROR_SIGNALS.search(body):
        return ERROR_BLOCK_TAG

    return CODE_BLOCK_TAG

print("Fenced block classifier ready")


# =========================
# QUOTE ENGINE
# =========================


def _line_quote_level(line):
    m = re.match(r'^\s*(>[>\s]*)', line)
    return m.group(1).count('>') if m else 0


def _strip_all_quote_prefixes(line):
    s = line
    while re.match(r'^\s*>\s?', s):
        s = re.sub(r'^\s*>\s?', '', s)
    return s.rstrip()


def _is_low_value_quote(text):
    if not isinstance(text, str):
        return True
    t = re.sub(r'\s+', ' ', text).strip()
    if len(t) < 8:
        return True
    if re.fullmatch(r'[\W_]+', t):
        return True
    if re.search(r'(?i)gibberish|giberrish|random characters', t):
        return True
    if len(re.findall(r'[A-Za-zÀ-ÿ]', t)) < 3:
        return True
    return False


def _is_technical_quote(text):
    if not isinstance(text, str):
        return False
    technical_tags = [CODE_BLOCK_TAG, DIFF_BLOCK_TAG, ERROR_BLOCK_TAG]
    return any(tag in text for tag in technical_tags)


def _is_complex_quote(text):
    if not isinstance(text, str):
        return False
    return ('\n' in text) or _is_technical_quote(text)


def _render_quote_context(quote_text, response=None):
    if not isinstance(quote_text, str) or not quote_text.strip():
        return response or ""
    quote_text = quote_text.strip()
    response = "" if response is None else response.strip()
    if _is_complex_quote(quote_text):
        prefix = f"based on the quotation:\n{quote_text}"
    else:
        q = "'" if '"' in quote_text else '"'
        prefix = f"based on the quotation: {q}{quote_text}{q}"
    if response:
        return f"{prefix}\n\n{response}"
    return prefix


def _split_quoted_segments(seg_lines):
    segments = []
    current = []
    in_fenced = False

    for raw in seg_lines:
        line = raw.rstrip("\n")
        content = re.sub(r'^\s*>\s?', '', line)

        if content.strip().startswith("```"):
            if not in_fenced:
                if current:
                    segments.append(("text", current))
                    current = []
                in_fenced = True
                current.append(content)
            else:
                current.append(content)
                segments.append(("fenced", current))
                current = []
                in_fenced = False
            continue

        current.append(content)

    if current:
        segments.append(("fenced" if in_fenced else "text", current))

    return segments


def _process_quote_block(seg_lines):
    parts = []
    buffer = []

    def flush_buffer():
        nonlocal buffer, parts
        if not buffer:
            return

        cleaned = []
        for line in buffer:
            content = re.sub(r'^\s*(?:>\s*)+', '', line.rstrip("\n")).strip()

            # conservar separación de párrafos
            if not content:
                if cleaned and cleaned[-1] != "":
                    cleaned.append("")
                continue

            # NO eliminar aquí citas "low value"
            cleaned.append(content)

        # quitar vacíos al inicio/fin, pero conservar vacíos internos
        while cleaned and cleaned[0] == "":
            cleaned.pop(0)
        while cleaned and cleaned[-1] == "":
            cleaned.pop()

        if cleaned:
            parts.append("\n".join(cleaned))

        buffer = []

    i = 0
    while i < len(seg_lines):
        raw = seg_lines[i].rstrip("\n")
        stripped = re.sub(r'^\s*(?:>\s*)+', '', raw)

        # detectar fenced block citado
        if stripped.strip().startswith("```"):
            flush_buffer()

            fenced_lines = [stripped]
            i += 1
            while i < len(seg_lines):
                nxt = seg_lines[i].rstrip("\n")
                nxt_stripped = re.sub(r'^\s*(?:>\s*)+', '', nxt)
                fenced_lines.append(nxt_stripped)
                if nxt_stripped.strip().startswith("```"):
                    break
                i += 1

            block_text = "\n".join(fenced_lines).strip()
            cleaned_block = replace_fenced_blocks(block_text).strip()
            if cleaned_block:
                parts.append(cleaned_block)

            i += 1
            continue

        buffer.append(raw)
        i += 1

    flush_buffer()

    return "\n\n".join(p for p in parts if p).strip()

def _extract_quotes(text):
    """
    Reglas:
    - cada línea con >   -> based on the quotation: "..."
    - cada línea con >>+ -> (citing: "...")
    - líneas vacías citadas se conservan
    - menciones dentro de la cita se anonimizarán, pero no se reescriben discursivamente
    """
    if not isinstance(text, str) or not text.strip():
        return [], ""

    lines = text.splitlines()
    quote_list = []
    result_lines = []

    i = 0
    while i < len(lines):
        line = lines[i]

        # línea no citada
        if not re.match(r'^\s*>', line):
            result_lines.append(line)
            i += 1
            continue

        level = _line_quote_level(line)
        content = _strip_all_quote_prefixes(line).strip()

        # línea citada vacía
        if not content:
            result_lines.append("")
            i += 1
            continue

        # bloque fenced citado
        if content.startswith("```"):
            fenced_lines = [content]
            i += 1

            while i < len(lines):
                nxt = lines[i]
                if not re.match(r'^\s*>', nxt):
                    break

                nxt_content = _strip_all_quote_prefixes(nxt).rstrip()
                fenced_lines.append(nxt_content)

                if nxt_content.strip().startswith("```") and len(fenced_lines) > 1:
                    break

                i += 1

            block_text = "\n".join(fenced_lines).strip()

            if "replace_fenced_blocks" in globals():
                cleaned_block = replace_fenced_blocks(block_text).strip()
                if not cleaned_block:
                    cleaned_block = CODE_BLOCK_TAG
            else:
                cleaned_block = CODE_BLOCK_TAG

            if level == 1:
                result_lines.append(f"based on the quotation: {cleaned_block}")
            else:
                q = "'" if '"' in cleaned_block else '"'
                result_lines.append(f"(citing: {q}{cleaned_block}{q})")

            quote_list.append(cleaned_block)
            i += 1
            continue

        # línea citada normal
        content = _anonymize_inline_mentions(content)
        q = "'" if '"' in content else '"'

        if level == 1:
            result_lines.append(f"based on the quotation: {q}{content}{q}")
        else:
            result_lines.append(f"(citing: {q}{content}{q})")

        quote_list.append(content)
        i += 1

    return quote_list, "\n".join(result_lines).strip()


# =================
# TAGS
# =================

CODE_BLOCK_TAG = "/Code_block_attached."
DIFF_BLOCK_TAG = "/Diff_attached."
ERROR_BLOCK_TAG = "/Error_log_attached."

USER_MENTION_SINGLE = "A GitHub user is mentioned."
USER_MENTION_MULTI = "Multiple GitHub users are mentioned."
USER_REVIEW_SINGLE = "A GitHub user is asked to review or check this."
USER_REVIEW_MULTI = "Multiple GitHub users are asked to review or check this."

USER_LEADING_SINGLE = "A GitHub user is mentioned."
USER_LEADING_MULTI = "Multiple GitHub users are mentioned."

TEAM_MENTION_SINGLE = "A team from a GitHub organization is mentioned."
TEAM_MENTION_MULTI = "Multiple teams from GitHub organizations are mentioned."
TEAM_REVIEW_SINGLE = "A team from a GitHub organization is asked to review or check the comment."
TEAM_REVIEW_MULTI = "Multiple teams from GitHub organizations are asked to review or check the comment."

TEAM_LEADING_SINGLE = "A team from a GitHub organization is mentioned."
TEAM_LEADING_MULTI = "Multiple teams from GitHub organizations are mentioned."

# =================
# MENTION PATTERNS
# =================

TEAM_MENTION_RE = re.compile(r'(?<!\w)@([A-Za-z0-9_.-]+)/([A-Za-z0-9_.-]+)')
USER_MENTION_RE = re.compile(r'(?<!\w)@([A-Za-z0-9_.-]+)\b')

REVIEW_CUES_RE = re.compile(
    r'(?i)\b('
    r'please review|can you review|could you review|'
    r'please take a look|take a look|'
    r'can you check|could you check|check this|ptal'
    r')\b'
)

ONLY_USER_MENTIONS_RE = re.compile(r'^\s*(?:@[A-Za-z0-9_.-]+\s*)+$')
ONLY_TEAM_MENTIONS_RE = re.compile(r'^\s*(?:@[A-Za-z0-9_.-]+/[A-Za-z0-9_.-]+\s*)+$')

GREETING_USER_RE = re.compile(
    r'^\s*(Hi|Hello|Hey)\s+@([A-Za-z0-9_.-]+)\s*([.!?]?)\s*$',
    re.IGNORECASE
)
GREETING_MULTI_USER_RE = re.compile(
    r'^\s*(Hi|Hello|Hey)\s+(?:@[A-Za-z0-9_.-]+\s*){2,}([.!?]?)\s*$',
    re.IGNORECASE
)

GREETING_THANKS_MULTI_USER_RE = re.compile(
    r'^\s*(Hi|Hello|Hey)\s+((?:@[A-Za-z0-9_.-]+\s*(?:,|\band\b)?\s*){2,})([!?\.]*)\s*(.*)$',
    re.IGNORECASE | re.DOTALL
)

GREETING_THANKS_SINGLE_USER_RE = re.compile(
    r'^\s*(Hi|Hello|Hey)\s+(@[A-Za-z0-9_.-]+)([!?\.]*)\s*(.*)$',
    re.IGNORECASE | re.DOTALL
)

LEADING_USER_MENTIONS_RE = re.compile(
    r'^\s*((?:@[A-Za-z0-9_.-]+\s*)+)(.+)$',
    re.DOTALL
)
LEADING_TEAM_MENTIONS_RE = re.compile(
    r'^\s*((?:@[A-Za-z0-9_.-]+/[A-Za-z0-9_.-]+\s*)+)(.+)$',
    re.DOTALL
)

THANKS_SINGLE_USER_RE = re.compile(
    r'(?i)^\s*(thanks|thank you)[,\s]+@[A-Za-z0-9_.-]+\s*[.!?]?\s*$'
)

THANKS_MULTI_USERS_RE = re.compile(
    r'(?i)^\s*(thanks|thank you)[,\s]+(?:@[A-Za-z0-9_.-]+\s*(?:,|\band\b)?\s*){2,}[.!?]?\s*$'
)

THANKS_LEADING_MULTI_USERS_RE = re.compile(
    r'(?i)^\s*(thanks|thank you)[,\s]+((?:@[A-Za-z0-9_.-]+\s*(?:,|\band\b)?\s*){2,})([.!?]?\s*)(.*)$'
)

THANKS_LEADING_SINGLE_USER_RE = re.compile(
    r'(?i)^\s*(thanks|thank you)[,\s]+(@[A-Za-z0-9_.-]+)([.!?]?\s*)(.*)$'
)
MEETING_WITH_MULTI_USERS_RE = re.compile(
    r'(?i)^(.*?\b(?:meeting|call|discussion|sync|chat)\s+with)\s+((?:@[A-Za-z0-9_.-]+\s*(?:,|\band\b)?\s*){2,})(\s*[:.-]?\s*)(.*)$',
    re.DOTALL
)

MEETING_WITH_SINGLE_USER_RE = re.compile(
    r'(?i)^(.*?\b(?:meeting|call|discussion|sync|chat)\s+with)\s+(@[A-Za-z0-9_.-]+)(\s*[:.-]?\s*)(.*)$',
    re.DOTALL
)


QUOTE_PREFIX_RE = re.compile(r'^\s*((?:>\s*)+)(.*)$')


QUESTION_START_RE = re.compile(
    r'(?i)^\s*(would|could|can|should|do|does|is|are|what|why|how)\b'
)

REQUEST_START_RE = re.compile(
    r'(?i)^\s*('
    r'please\b|can you\b|could you\b|would you\b|'
    r'take a look\b|check\b|review\b|have a look\b'
    r')'
)

PROPOSAL_START_RE = re.compile(
    r'(?i)^\s*(would it make sense to|does it make sense to|maybe we should|perhaps we should|it might make sense to)\b'
)

THOUGHT_OPINION_RE = re.compile(
    r'(?i)^\s*(what do you think|thoughts\??)\s*$'
)

OPINION_THEN_FEEDBACK_RE = re.compile(
    r'(?is)^\s*(i think|i believe|in my opinion|imo)\b.*\b(what do you think\??|thoughts\??)\s*$'
)

ENDS_WITH_FEEDBACK_RE = re.compile(
    r'(?is)^(.*?)[,;:\s-]*(what do you think\??|thoughts\??)\s*$'
)

COMMIT_LINE_REFERENCE_RE = re.compile(
    r'(?i)\bLine\s+\d+\s+in\s+\[?/hash_reference\]?\([^)]+\)'
)
CC_USER_MENTION_RE = re.compile(
    r'(?i)(?<!\w)/?cc\s+@[A-Za-z0-9_.-]+(?:\s+as well)?'
)

APPROVAL_NOTIFICATION_RE = re.compile(
    r'(?is)\[APPROVALNOTIFIER\].*?\bAPPROVED\b(?:\*\*)?'
)

SUGGESTION_RE = re.compile(
    r'(?im)^\s*#{0,6}\s*suggestion\s*$'
)

# =================
# HELPERS
# =================

def _team_spans(text):
    return [m.span() for m in TEAM_MENTION_RE.finditer(text)] if isinstance(text, str) else []

def count_team_mentions(text):
    if not isinstance(text, str):
        return 0
    return len(list(TEAM_MENTION_RE.finditer(text)))

def count_user_mentions(text):
    if not isinstance(text, str):
        return 0

    spans = _team_spans(text)
    count = 0
    for m in USER_MENTION_RE.finditer(text):
        s, e = m.span()
        inside_team = any(s >= ts and e <= te for ts, te in spans)
        if not inside_team:
            count += 1
    return count

def _leading_user_mentions_info(text):
    if not isinstance(text, str):
        return 0, ""
    m = re.match(r'^\s*((?:@[A-Za-z0-9_.-]+\s*)+)(.*)$', text, re.DOTALL)
    if not m:
        return 0, text
    lead = m.group(1)
    rest = m.group(2)
    n = len(re.findall(r'@[A-Za-z0-9_.-]+', lead))
    return n, rest

def _normalize_after_prefix(text):
    text = re.sub(r'^[\s,;:.-]+', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def _extract_quote_info(line):
    """
    Returns
    -------
    depth : int
        Number of quote markers at the beginning of the line.
    content : str
        Line content without the leading Markdown quote markers.
    """
    if not isinstance(line, str):
        return 0, line

    m = QUOTE_PREFIX_RE.match(line)
    if not m:
        return 0, line

    prefix = m.group(1)
    content = m.group(2).strip()
    depth = prefix.count(">")
    return depth, content


def _anonymize_inline_mentions(text):
    text = COMMIT_LINE_REFERENCE_RE.sub("/commit_line_reference", text)
    text = TEAM_MENTION_RE.sub("a team from a GitHub organization", text)
    text = USER_MENTION_RE.sub("a GitHub user", text)
    return text

def clean_approval_notifications(text):
    if not isinstance(text, str):
        return ""

    return APPROVAL_NOTIFICATION_RE.sub(
        "/approval_notification",
        text
    )


def clean_code_suggestions(text):
    if not isinstance(text, str):
        return ""

    return SUGGESTION_RE.sub(
        "/code_suggestion",
        text
    )


def clean_cc_user_mentions(text):
    if not isinstance(text, str):
        return ""

    return CC_USER_MENTION_RE.sub(
        USER_MENTION_SINGLE,
        text
    )

def _rewrite_greeting_with_mentions(greeting, is_multi, trailing_text):
    greeting = greeting.capitalize().strip()
    trailing_text = trailing_text.strip()

    mention_text = USER_MENTION_MULTI if is_multi else USER_MENTION_SINGLE

    if not trailing_text:
        return f"{greeting}. {mention_text}"

    # normalizar texto después del greeting
    trailing_text = _anonymize_inline_mentions(trailing_text)

    # caso: agradecimiento
    if re.match(r'(?i)^thank you\b', trailing_text):
        return f"{greeting}. {mention_text} The comment thanks them for their work so far."

    # caso general
    return f"{greeting}. {mention_text} {trailing_text}"

# =================
# REWRITERS
# =================

def _rewrite_user_leading_context(rest, lead_n, has_review):
    rest = _normalize_after_prefix(rest)

    if not rest:
        return USER_LEADING_SINGLE if lead_n == 1 else USER_LEADING_MULTI

    rest_clean = rest.strip()
    rest_lower = rest_clean.lower()

       # only asking opinion
    if THOUGHT_OPINION_RE.fullmatch(rest_clean):
        return (
            f"{USER_LEADING_SINGLE} The message asks for that user's opinion."
            if lead_n == 1
            else f"{USER_LEADING_MULTI} The message asks for their opinion."
        )

    # opinion + asks back
    if OPINION_THEN_FEEDBACK_RE.match(rest_clean):
        m = ENDS_WITH_FEEDBACK_RE.match(rest_clean)
        main_part = m.group(1).strip() if m else rest_clean
        return (
            f"{USER_LEADING_SINGLE} {main_part} The message then asks for that user's opinion."
            if lead_n == 1
            else f"{USER_LEADING_MULTI} {main_part} The message then asks for their opinion."
        )

    # explicit review/check
    if has_review:
        cleaned = re.sub(
            r'(?i)^(please review|can you review|could you review|please take a look|take a look|can you check|could you check|check this|ptal)\b[\s,:-]*',
            '',
            rest_clean
        ).strip()

        return (
            USER_REVIEW_SINGLE + (f" {cleaned}" if cleaned else "")
            if lead_n == 1
            else USER_REVIEW_MULTI + (f" {cleaned}" if cleaned else "")
        )

    # proposal
    if PROPOSAL_START_RE.match(rest_clean):
        if rest_lower.startswith("would it make sense to"):
            tail = rest_clean[len("would it make sense to"):].strip()
            return (
                f"{USER_LEADING_SINGLE} It asks whether it would make sense to {tail}"
                if lead_n == 1
                else f"{USER_LEADING_MULTI} It asks whether it would make sense to {tail}"
            )

        return (
            f"{USER_LEADING_SINGLE} It asks whether {rest_clean}"
            if lead_n == 1
            else f"{USER_LEADING_MULTI} It asks whether {rest_clean}"
        )

    # question
    if QUESTION_START_RE.match(rest_clean):
        return (
            f"{USER_LEADING_SINGLE} It asks: {rest_clean}"
            if lead_n == 1
            else f"{USER_LEADING_MULTI} It asks: {rest_clean}"
        )

    # request
    if REQUEST_START_RE.match(rest_clean):
        cleaned_request_rest = re.sub(
            r'(?i)^(please|can you|could you|would you|take a look|check|review|have a look)\b[\s,:-]*',
            '',
            rest_clean
        ).strip()

        return (
            USER_REVIEW_SINGLE + (f" {cleaned_request_rest}" if cleaned_request_rest else "")
            if lead_n == 1
            else USER_REVIEW_MULTI + (f" {cleaned_request_rest}" if cleaned_request_rest else "")
        )

    # generic directed statement
    return (
        f"{USER_LEADING_SINGLE} {rest_clean}"
        if lead_n == 1
        else f"{USER_LEADING_MULTI} {rest_clean}"
    )

def _rewrite_team_leading_context(rest, team_count, has_review):
    rest = _normalize_after_prefix(rest)

    if not rest:
        return TEAM_LEADING_SINGLE if team_count == 1 else TEAM_LEADING_MULTI

    rest_clean = rest.strip()

    if has_review:
        return (
            f"{TEAM_REVIEW_SINGLE} {rest_clean}"
            if team_count == 1
            else f"{TEAM_REVIEW_MULTI} {rest_clean}"
        )

    if PROPOSAL_START_RE.match(rest_clean):
        return (
            f"{TEAM_LEADING_SINGLE} It asks whether {rest_clean}"
            if team_count == 1
            else f"{TEAM_LEADING_MULTI} It asks whether {rest_clean}"
        )

    if QUESTION_START_RE.match(rest_clean):
        return (
            f"{TEAM_LEADING_SINGLE} It asks: {rest_clean}"
            if team_count == 1
            else f"{TEAM_LEADING_MULTI} It asks: {rest_clean}"
        )

    if REQUEST_START_RE.match(rest_clean):
        return (
            f"{TEAM_REVIEW_SINGLE} {rest_clean}"
            if team_count == 1
            else f"{TEAM_REVIEW_MULTI} {rest_clean}"
        )

    return (
        f"{TEAM_LEADING_SINGLE} {rest_clean}"
        if team_count == 1
        else f"{TEAM_LEADING_MULTI} {rest_clean}"
    )
def _clean_quote_line(line):
    depth, content = _extract_quote_info(line)

    if depth == 0:
        return line

    if not content:
        return ""

    quote_intro = "Based on the quotation, "


    # solo menciones de equipo
    if ONLY_TEAM_MENTIONS_RE.fullmatch(stripped):
        team_count = count_team_mentions(stripped)
        return TEAM_MENTION_SINGLE if team_count == 1 else TEAM_MENTION_MULTI


    # solo menciones de usuario
    if ONLY_USER_MENTIONS_RE.fullmatch(stripped):
        user_count = count_user_mentions(stripped)
        return USER_MENTION_SINGLE if user_count == 1 else USER_MENTION_MULTI

    team_count = count_team_mentions(stripped)
    user_count = count_user_mentions(stripped)

    # cita que comienza con mención de equipo
    m_team = LEADING_TEAM_MENTIONS_RE.match(content)
    if m_team and count_team_mentions(content) > 0 and count_user_mentions(content) == 0:
        rest = _normalize_after_prefix(m_team.group(2))
        if rest:
            return quote_intro + f"{TEAM_LEADING_SINGLE} {rest}"
        return quote_intro + TEAM_MENTION_SINGLE

    # cita que comienza con mención de usuario
    m_user = LEADING_USER_MENTIONS_RE.match(content)
    if m_user and count_user_mentions(content) > 0 and count_team_mentions(content) == 0:
        lead_n, rest = _leading_user_mentions_info(content)
        rewritten = _rewrite_user_leading_context(
            rest,
            lead_n,
            bool(REVIEW_CUES_RE.search(content))
        )
        return quote_intro + rewritten

    # cita general
    content = _anonymize_inline_mentions(content)
    return quote_intro + content



# =================
# MAIN
# =================

def clean_mentions(text):
    if not isinstance(text, str) or not text.strip():
        return ""

    stripped = text.strip()

    depth, _ = _extract_quote_info(stripped)
    if depth > 0:
        return _clean_quote_line(stripped)

    # isolated team mention(s)
    if ONLY_TEAM_MENTIONS_RE.fullmatch(stripped):
        team_count = count_team_mentions(stripped)
        return TEAM_MENTION_SINGLE if team_count == 1 else TEAM_MENTION_MULTI

    # isolated user mention(s)
    if ONLY_USER_MENTIONS_RE.fullmatch(stripped):
        user_count = count_user_mentions(stripped)
        return USER_MENTION_SINGLE if user_count == 1 else USER_MENTION_MULTI

    # thanks + single user
    if THANKS_SINGLE_USER_RE.fullmatch(stripped):
        return "The comment thanks a GitHub user."

    # thanks + multiple users
    if THANKS_MULTI_USERS_RE.fullmatch(stripped):
        return "The comment thanks multiple GitHub users."

    # thanks + multiple users at the beginning, followed by more content
    m_thanks_multi = THANKS_LEADING_MULTI_USERS_RE.match(stripped)
    if m_thanks_multi:
        rest = m_thanks_multi.group(4).strip()
        if rest:
            return f"The comment thanks multiple GitHub users. {rest}"
        return "The comment thanks multiple GitHub users."

    # thanks + single user at the beginning, followed by more content
    m_thanks_single = THANKS_LEADING_SINGLE_USER_RE.match(stripped)
    if m_thanks_single:
        rest = m_thanks_single.group(4).strip()
        if rest:
            return f"The comment thanks a GitHub user. {rest}"
        return "The comment thanks a GitHub user."

    # meeting/discussion with multiple users
    m_meeting_multi = MEETING_WITH_MULTI_USERS_RE.match(stripped)
    if m_meeting_multi:
        prefix = m_meeting_multi.group(1).strip()
        sep = m_meeting_multi.group(3) or ""
        rest = m_meeting_multi.group(4).strip()

        rewritten = f"{prefix} multiple GitHub users"
        if sep:
            rewritten += sep.rstrip()
        if rest:
            rewritten += f" {rest}"
        return rewritten

    # meeting/discussion with a single user
    m_meeting_single = MEETING_WITH_SINGLE_USER_RE.match(stripped)
    if m_meeting_single:
        prefix = m_meeting_single.group(1).strip()
        sep = m_meeting_single.group(3) or ""
        rest = m_meeting_single.group(4).strip()

        rewritten = f"{prefix} a GitHub user"
        if sep:
            rewritten += sep.rstrip()
        if rest:
            rewritten += f" {rest}"
        return rewritten

    # greeting + multiple users + extra text
    m_greet_thanks_multi = GREETING_THANKS_MULTI_USER_RE.match(stripped)
    if m_greet_thanks_multi:
        greeting = m_greet_thanks_multi.group(1)
        trailing_text = m_greet_thanks_multi.group(4)
        return _rewrite_greeting_with_mentions(greeting, is_multi=True, trailing_text=trailing_text)

    # greeting + single user + extra text
    m_greet_thanks_single = GREETING_THANKS_SINGLE_USER_RE.match(stripped)
    if m_greet_thanks_single:
        greeting = m_greet_thanks_single.group(1)
        trailing_text = m_greet_thanks_single.group(4)
        return _rewrite_greeting_with_mentions(greeting, is_multi=False, trailing_text=trailing_text)

    # greeting + single user
    m_greet_single = GREETING_USER_RE.fullmatch(stripped)
    if m_greet_single:
        greeting = m_greet_single.group(1).capitalize()
        return f"{greeting}. {USER_MENTION_SINGLE}"

    # greeting + multiple users
    m_greet_multi = GREETING_MULTI_USER_RE.fullmatch(stripped)
    if m_greet_multi:
        greeting = m_greet_multi.group(1).capitalize()
        return f"{greeting}. {USER_MENTION_MULTI}"

    team_count = count_team_mentions(stripped)
    user_count = count_user_mentions(stripped)
    has_review = bool(REVIEW_CUES_RE.search(stripped))


    # leading team mention(s)
    m_team = LEADING_TEAM_MENTIONS_RE.match(stripped)
    if m_team and team_count > 0 and user_count == 0:
        return _rewrite_team_leading_context(m_team.group(2), team_count, has_review)

    # leading user mention(s)
    m_user = LEADING_USER_MENTIONS_RE.match(stripped)
    if m_user and user_count > 0 and team_count == 0:
        lead_n, rest_raw = _leading_user_mentions_info(stripped)
        return _rewrite_user_leading_context(rest_raw, lead_n, has_review)

    # inline mentions
    return _anonymize_inline_mentions(stripped)

def clean_mentions_by_paragraph(text):
    if not isinstance(text, str) or not text.strip():
        return ""

    lines = text.splitlines()
    cleaned_lines = []

    for line in lines:
        if not line.strip():
            cleaned_lines.append("")
            continue

        cleaned = clean_mentions(line)
        if cleaned:
            cleaned_lines.append(cleaned)

    return "\n".join(cleaned_lines).strip()

print("Mention cleaner ready")

def normalize_comment_text(text):
    if not isinstance(text, str) or not text.strip():
        return ""

    # primero citas
    _, text = _extract_quotes(text)

    # luego menciones y párrafos
    text = clean_mentions_by_paragraph(text)

    return text

    final_text = normalize_comment_text(text)

# =========================
# WORKFLOW COMMANDS
# =========================

SLASH_TAXONOMY = {
    "/assign": "/workflow_coordination",
    "/unassign": "/workflow_coordination",
    "/cc": "/workflow_coordination",
    "/sig": "/workflow_coordination",
    "/wg": "/workflow_coordination",
    "/label": "/workflow_coordination",
    "/remove-label": "/workflow_coordination",
    "/milestone": "/workflow_coordination",
    "/priority": "/workflow_coordination",
    "/remove-priority": "/workflow_coordination",
    "/kind": "/workflow_coordination",
    "/remove-kind": "/workflow_coordination",
    "/triage": "/workflow_coordination",

    "/test": "/workflow_ci",
    "/retest": "/workflow_ci",
    "/retest-required": "/workflow_ci",
    "/ok-to-test": "/workflow_ci",
    "/skip": "/workflow_ci",

    "/approve": "/workflow_review",
    "/lgtm": "/workflow_review",

    "/hold": "/workflow_blocking",
    "/close": "/workflow_blocking",
    "/reopen": "/workflow_blocking",
}

KNOWN_SLASH_COMMANDS = (
    "assign|unassign|cc|sig|wg|label|remove-label|milestone|priority|remove-priority|"
    "kind|remove-kind|triage|test|retest|retest-required|ok-to-test|skip|approve|"
    "lgtm|hold|close|reopen"
)

SLASH_CMD_RE = re.compile(rf'(?m)(?<!\S)(/(?:{KNOWN_SLASH_COMMANDS}))\b')
SLASH_ONLY_RE = re.compile(rf'^\s*(?:(?:/(?:{KNOWN_SLASH_COMMANDS}))(?:\s+[@\w.\-/]+)*\s*)+$')

PLAIN_CC_ONLY_RE = re.compile(r'(?im)^\s*cc\s+@?[\w.\-]+(?:\s+@?[\w.\-]+)*\s*$')
PLAIN_CC_LEADING_RE = re.compile(r'(?im)^\s*cc\s+@?[\w.\-]+(?:\s+@?[\w.\-]+)*(.*)$')

def extract_slash_commands(text):
    if not isinstance(text, str):
        return []
    return SLASH_CMD_RE.findall(text)

def classify_slash_commands(text):
    cmds = extract_slash_commands(text)
    types = []
    for cmd in cmds:
        types.append(SLASH_TAXONOMY.get(cmd, "/workflow_other"))
    return sorted(set(types))

def is_slash_only(text):
    if not isinstance(text, str) or not text.strip():
        return False
    return bool(SLASH_ONLY_RE.fullmatch(text.strip()))

def replace_plain_cc(text):
    if not isinstance(text, str) or not text.strip():
        return ""
    stripped = text.strip()
    if PLAIN_CC_ONLY_RE.fullmatch(stripped):
        return "/workflow_coordination"
    m = PLAIN_CC_LEADING_RE.match(stripped)
    if m:
        rest = re.sub(r'^[,;:.-]+\s*', '', m.group(1).strip())
        return rest if rest else "/workflow_coordination"
    return text

def clean_workflow_commands(text):
    if not isinstance(text, str) or not text.strip():
        return ""
    stripped = text.strip()
    cmd_types = classify_slash_commands(stripped)

    stripped = replace_plain_cc(stripped)
    if stripped == "/workflow_coordination":
        return stripped

    if not cmd_types:
        return stripped

    if is_slash_only(text):
        return " ".join(cmd_types)

    cleaned = re.sub(
    rf'(?<!\S)/(?:{KNOWN_SLASH_COMMANDS})\b',
    '',
    stripped
)
    cleaned = re.sub(r'[ \t]+', ' ', cleaned).strip()
    cleaned = re.sub(r'^[,;:.-]+\s*', '', cleaned)
    cleaned = re.sub(r'\s+([,;:.!?])', r'\1', cleaned)
    # mixed comments: preserve message, do not prefix workflow tags aggressively
    return cleaned if cleaned else " ".join(cmd_types)

print("Workflow command cleaner ready")
# =========================
# BOTS AND AUTOMATED MESSAGES
# =========================

BOT_PATTERN = re.compile(
    r'(?i)'
    r'\[bot\]|'
    r'ci[\-_]robot|triage[\-_]robot|'
    r'dependabot|renovate|greenkeeper|'
    r'codecov|coveralls|mergify|'
    r'stale\[bot\]|'
    r'allcontributors|'
    r'nodejs-github-bot|'
    r'welcome[\-_]bot|onboarding[\-_]bot|'
    r'approval[\-_]bot|review[\-_]bot|'
    r'flaky[\-_]bot|retest[\-_]bot'
)

TRIAGE_BOT_RE = re.compile(r'(?i)triage[\-_]robot|triage[\-_]bot')
APPROVAL_BOT_RE = re.compile(r'(?i)approval[\-_]bot|review[\-_]bot|mergify')
RETEST_FLAKY_BOT_RE = re.compile(r'(?i)flaky[\-_]bot|retest[\-_]bot|ci[\-_]robot|codecov|coveralls')
ONBOARDING_BOT_RE = re.compile(r'(?i)welcome[\-_]bot|onboarding[\-_]bot|allcontributors')

def classify_bot_or_automation(author, text=""):
    author = "" if author is None else str(author)
    text = "" if text is None else str(text)
    hay = f"{author} {text}"

    if not BOT_PATTERN.search(hay):
        return ""

    if TRIAGE_BOT_RE.search(hay):
        return "triage_bot"
    if APPROVAL_BOT_RE.search(hay):
        return "approval_notifier"
    if RETEST_FLAKY_BOT_RE.search(hay):
        return "retest_flaky_bot"
    if ONBOARDING_BOT_RE.search(hay):
        return "onboarding_bot"
    return "other_bot"

def is_bot_or_automation(author, text=""):
    return bool(classify_bot_or_automation(author, text))

def _looks_like_pure_ci_or_coverage_message(text):
    if not isinstance(text, str):
        return False
    lines = [ln.strip() for ln in str(text).splitlines() if ln.strip()]
    if not lines:
        return False
    return all(
        CI_STATUS_LINE_RE.match(ln) or
        CODECOV_STATUS_LINE_RE.match(ln) or
        COVERALLS_STATUS_LINE_RE.match(ln)
        for ln in lines
    )

def clean_bot_messages(text, author=""):
    if not isinstance(text, str):
        return ""
    bot_type = classify_bot_or_automation(author, text)
    if not bot_type:
        return text
    stripped = text.strip()

    # Let CI/coverage layers handle pure automated report lines.
    if _looks_like_pure_ci_or_coverage_message(stripped):
        return stripped

    return f"{BOT_TAG} {stripped}".strip() if stripped else BOT_TAG

print("Bot and automation rules ready")



# ---

# Integración opcional con Google Drive (solo si el notebook corre en Colab)
try:
    from google.colab import drive
    drive.mount('/content/drive')
    print("Google Drive montado correctamente.")
except Exception:
    print("Google Colab o Google Drive no está disponible en este entorno. Se continúa sin montaje.")


# ---

examples_workflow = [
    "/assign @alice",
    "/approve",
    "Done. /assign @alice",
    "Please /test this again.",
    "I think we should /assign @alice later.",
    "/cc @alice @bob",
]

for text in examples_workflow:
    print("=" * 80)
    print("ORIGINAL:", text)
    print("CLEANED :", clean_workflow_commands(text))

# ---

# Función que reemplaza fenced blocks normales y citados
QUOTED_FENCED_BLOCK_RE = re.compile(
    r"(^\s*>\s*```([A-Za-z0-9_+\-]*)[ \t]*\n(?:^\s*>\s?.*\n)*?^\s*>\s*```)",
    re.MULTILINE
)

def _strip_quote_prefixes_from_block(block_text):
    lines = block_text.splitlines()
    cleaned = [re.sub(r"^\s*>\s?", "", line) for line in lines]
    return "\n".join(cleaned)

def _replace_single_fenced_block(block_text):
    match = FENCED_BLOCK_RE.search(block_text)
    if not match:
        return block_text

    lang = (match.group(1) or "").strip()
    body = match.group(2) or ""
    label = classify_fenced_block(lang, body)
    return f"\n{label}\n"

def replace_fenced_blocks(text):
    if not isinstance(text, str):
        return ""

    def quoted_repl(match):
        quoted_block = match.group(1)
        cleaned_block = _strip_quote_prefixes_from_block(quoted_block)
        return _replace_single_fenced_block(cleaned_block)

    def fenced_repl(match):
        lang = (match.group(1) or "").strip()
        body = match.group(2) or ""
        label = classify_fenced_block(lang, body)
        return f"\n{label}\n"

    text = QUOTED_FENCED_BLOCK_RE.sub(quoted_repl, text)
    text = FENCED_BLOCK_RE.sub(fenced_repl, text)
    return text

print("Fenced block replacement ready")


# ---

# Detect diff outside fenced blocks

def replace_diff_structures(text):
    if not isinstance(text, str):
        return ""

    diff_signals = 0
    if DIFF_GIT_RE.search(text):
        diff_signals += 1
    if DIFF_HUNK_RE.search(text):
        diff_signals += 1
    if DIFF_FILE_OLD_RE.search(text):
        diff_signals += 1
    if DIFF_FILE_NEW_RE.search(text):
        diff_signals += 1
    if DIFF_INDEX_RE.search(text):
        diff_signals += 1

    if diff_signals >= 2:
        return DIFF_BLOCK_TAG

    return text

print("Plain diff replacement ready")


# ---

# Reemplazar error logs en texto plano

def replace_error_logs(text):
    if not isinstance(text, str):
        return ""

    # Encapsulate only when the text looks like a real error log or traceback
    if ERROR_LOG_STRONG_RE.search(text):
        return ERROR_BLOCK_TAG

    return text

print("Plain error log replacement ready")


# ---

# Emoticons and emojis
import unicodedata

# =========================
# EMOJI / EMOTICON NORMALIZATION
# =========================

EMO_RISA_TAG = "/emoji_risa"
EMO_SONRISA_TAG = "/emoji_sonrisa"
EMO_AMOR_TAG = "/emoji_amor"
EMO_CELEBRA_TAG = "/emoji_celebra"
EMO_APROBACION_TAG = "/emoji_aprobacion"
EMO_RECHAZA_TAG = "/emoji_rechaza"
EMO_TRISTEZA_TAG = "/emoji_tristeza"
EMO_RABIA_TAG = "/emoji_rabia"
EMO_ASCO_TAG = "/emoji_asco"
EMO_MIEDO_TAG = "/emoji_miedo"
EMO_SORPRESA_TAG = "/emoji_sorpresa"
EMO_CONFUSION_TAG = "/emoji_confusion"
EMO_SARCASMO_TAG = "/emoji_sarcasmo"
EMO_CALMA_TAG = "/emoji_calma"
EMO_INTENSO_TAG = "/emoji_intenso"
EMO_ATENCION_TAG = "/emoji_atencion"
EMO_ESPERANZA_TAG = "/emoji_esperanza"
FLAG_TAG = "/emoji_flag_reference"
EMO_OTHER_TAG = "/emoji_other"

ASCII_EMOTICON_MAP = {
    r'(:-D|:D|=D|xD|XD|XDD|:\'D)': EMO_RISA_TAG,
    r'(:-\)|:\)|=\)|\^\^)': EMO_SONRISA_TAG,
    r'(<3)': EMO_AMOR_TAG,
    r'(:-\(|:\(|:\'\(|T_T|;_;)': EMO_TRISTEZA_TAG,
    r'(>:\(|D:<)': EMO_RABIA_TAG,
    r'(:-O|:O|O_O|o_O)': EMO_SORPRESA_TAG,
    r'(:-/|:/|:\\\\|o\.O|O\.o)': EMO_CONFUSION_TAG,
    r'(;-\)|;\)|:-P|:P|:-p|:p|XP)': EMO_SARCASMO_TAG,
}
GITHUB_EMOJI_SHORTCODE_MAP = {
    ":eyes:": EMO_ATENCION_TAG,
}

UNICODE_EMOJI_MAP = {
    # risa / humor
    "😂": EMO_RISA_TAG, "🤣": EMO_RISA_TAG, "😄": EMO_RISA_TAG,
    "😆": EMO_RISA_TAG, "😁": EMO_RISA_TAG, "😹": EMO_RISA_TAG,

    # sonrisa / agrado suave
    "🙂": EMO_SONRISA_TAG, "😊": EMO_SONRISA_TAG, "☺️": EMO_SONRISA_TAG,
    "😇": EMO_SONRISA_TAG,

    # amor / afecto
    "😍": EMO_AMOR_TAG, "🥰": EMO_AMOR_TAG, "😘": EMO_AMOR_TAG,
    "💚": EMO_AMOR_TAG, "❤️": EMO_AMOR_TAG, "❤": EMO_AMOR_TAG,

    # celebración / logro
    "🎉": EMO_CELEBRA_TAG, "🎊": EMO_CELEBRA_TAG, "🏅": EMO_CELEBRA_TAG,
    "🎁": EMO_CELEBRA_TAG, "👏": EMO_CELEBRA_TAG, "✨": EMO_CELEBRA_TAG,

    # aprobación
    "👍": EMO_APROBACION_TAG, "✅": EMO_APROBACION_TAG,

    # rechazo / negación
    "👎": EMO_RECHAZA_TAG, "❌": EMO_RECHAZA_TAG,

    # tristeza / decepción
    "😢": EMO_TRISTEZA_TAG, "😞": EMO_TRISTEZA_TAG,

    # rabia / frustración
    "😡": EMO_RABIA_TAG,

    # asco / repulsión
    "👹": EMO_ASCO_TAG,

    # confusión / duda
    "🤔": EMO_CONFUSION_TAG, "😕": EMO_CONFUSION_TAG, "😐": EMO_CONFUSION_TAG,
    "🤦": EMO_CONFUSION_TAG, "🤦‍♀️": EMO_CONFUSION_TAG,

    # calma / alivio
    "😌": EMO_CALMA_TAG, "🙇": EMO_CALMA_TAG,

    # sarcasmo / juego
    "😉": EMO_SARCASMO_TAG, "😅": EMO_SARCASMO_TAG,

    # atención / observación
    "👀": EMO_ATENCION_TAG,

    # gesto social / esperanza
    "👋": EMO_APROBACION_TAG, "🙏": EMO_ESPERANZA_TAG,
}

FLAG_RE = re.compile(r'[\U0001F1E6-\U0001F1FF]{2}')

EMOJI_CLUSTER_RE = re.compile(
    r'('
    r'(?:[\U0001F1E6-\U0001F1FF]{2})'
    r'|(?:['
    r'\U0001F300-\U0001FAFF'
    r'\U00002600-\U000027BF'
    r']'
    r'(?:\uFE0F)?'
    r'(?:[\U0001F3FB-\U0001F3FF])?'
    r'(?:\u200D['
    r'\U0001F300-\U0001FAFF'
    r'\U00002600-\U000027BF'
    r'](?:\uFE0F)?(?:[\U0001F3FB-\U0001F3FF])?)*'
    r')'
    r')'
)

def _classify_emoji_cluster(cluster):
    if FLAG_RE.fullmatch(cluster):
        return FLAG_TAG
    if cluster in UNICODE_EMOJI_MAP:
        return UNICODE_EMOJI_MAP[cluster]

    found_tags = []
    for emo, tag in UNICODE_EMOJI_MAP.items():
        if emo in cluster:
            found_tags.append(tag)

    if found_tags:
        priority = [
            EMO_AMOR_TAG, EMO_RISA_TAG, EMO_CELEBRA_TAG, EMO_APROBACION_TAG,
            EMO_RECHAZA_TAG, EMO_TRISTEZA_TAG, EMO_RABIA_TAG, EMO_ASCO_TAG,
            EMO_MIEDO_TAG, EMO_SORPRESA_TAG, EMO_CONFUSION_TAG, EMO_SARCASMO_TAG,
            EMO_CALMA_TAG, EMO_INTENSO_TAG, EMO_ATENCION_TAG, EMO_ESPERANZA_TAG
        ]
        for candidate in priority:
            if candidate in found_tags:
                return candidate

    return EMO_OTHER_TAG

def replace_emoticons_and_emojis(text):
    if not isinstance(text, str):
        return ""

    text = unicodedata.normalize("NFKC", text)

    # Shortcodes de GitHub, por ejemplo :eyes:
    for shortcode, tag in GITHUB_EMOJI_SHORTCODE_MAP.items():
        text = re.sub(
            re.escape(shortcode),
            f" {tag} ",
            text,
            flags=re.IGNORECASE
        )

    # Emojis / símbolos Unicode
    for emo in sorted(UNICODE_EMOJI_MAP, key=len, reverse=True):
        text = text.replace(
            emo,
            f" {UNICODE_EMOJI_MAP[emo]} "
        )

    # Emoticones ASCII
    for pattern, tag in ASCII_EMOTICON_MAP.items():
        text = re.sub(
            pattern,
            f" {tag} ",
            text
        )

    def emoji_repl(match):
        cluster = match.group(0)
        return f" {_classify_emoji_cluster(cluster)} "

    text = EMOJI_CLUSTER_RE.sub(emoji_repl, text)

    # Normalizar espacios, tabulaciones y saltos de línea
    text = re.sub(r"\s+", " ", text)

    return text.strip()


def has_emoticon_or_emoji(text):
    if not isinstance(text, str):
        return False

    norm = unicodedata.normalize("NFKC", text)

    # Shortcodes de GitHub
    for shortcode in GITHUB_EMOJI_SHORTCODE_MAP:
        if re.search(
            re.escape(shortcode),
            norm,
            flags=re.IGNORECASE
        ):
            return True

    # Emoticones ASCII
    for pattern in ASCII_EMOTICON_MAP:
        if re.search(pattern, norm):
            return True

    # Emojis Unicode
    return bool(EMOJI_CLUSTER_RE.search(norm))


def count_emoticon_or_emoji(text):
    if not isinstance(text, str):
        return 0

    norm = unicodedata.normalize("NFKC", text)
    count = 0

    # Shortcodes de GitHub
    for shortcode in GITHUB_EMOJI_SHORTCODE_MAP:
        count += len(
            re.findall(
                re.escape(shortcode),
                norm,
                flags=re.IGNORECASE
            )
        )

    # Emoticones ASCII
    for pattern in ASCII_EMOTICON_MAP:
        count += len(re.findall(pattern, norm))

    # Emojis Unicode
    count += len(EMOJI_CLUSTER_RE.findall(norm))

    return count


def detect_emotion_types(text):
    if not isinstance(text, str):
        return []

    norm = unicodedata.normalize("NFKC", text)
    found = set()

    # Shortcodes de GitHub
    for shortcode, tag in GITHUB_EMOJI_SHORTCODE_MAP.items():
        if re.search(
            re.escape(shortcode),
            norm,
            flags=re.IGNORECASE
        ):
            found.add(tag)

    # Emoticones ASCII
    for pattern, tag in ASCII_EMOTICON_MAP.items():
        if re.search(pattern, norm):
            found.add(tag)

    # Emojis Unicode
    for match in EMOJI_CLUSTER_RE.finditer(norm):
        found.add(
            _classify_emoji_cluster(match.group(0))
        )

    return sorted(found)

# ---


# Technical references

ISSUE_REF_TAG = "/issue_reference"
REPO_ISSUE_REF_TAG = "/repo_issue_reference"
HASH_TAG = "/hash_reference"
LOG_TAG = "/log_reference"
BENCHMARK_TAG = "/benchmark_reference"

REPO_ISSUE_REF_RE = re.compile(r'\b[\w.-]+/[\w.-]+#\d+\b')
ISSUE_REF_RE = re.compile(r'(?<![\w/])#\d+\b')

FULL_HASH_RE = re.compile(r'\b[0-9a-f]{40}\b', re.IGNORECASE)
SHORT_HASH_RE = re.compile(r'\b[0-9a-f]{7,12}\b', re.IGNORECASE)
GIT_TREE_HASH_RE = re.compile(r'(?im)^\s*index\s+[0-9a-f]+\.\.[0-9a-f]+\s+\d+\s*$')

LOG_LINE_RE = re.compile(
    r'''(?im)^\s*(?:\[[^\]]+\]\s*)?(?:INFO|DEBUG|WARN|WARNING|ERROR|TRACE|FATAL)\b.*$
    |^\s*at\s+.+$
    |^\s*File\s+"[^"]+",\s+line\s+\d+.*$
    |^\s*Caused by:.*$
    ''',
    re.VERBOSE
)

BENCHMARK_RE = re.compile(
    r'''(?im)^\s*.*(?:
        \b\d+(?:\.\d+)?\s*(?:ns/op|µs/op|us/op|ms/op|s/op)\b|
        \b\d+(?:\.\d+)?\s*(?:ops/sec|op/s|iter/s|iterations/s)\b|
        \bbenchmark\b|
        \bthroughput\b|
        \blatency\b
    ).*$''',
    re.VERBOSE
)

def has_issue_ref(text):
    if not isinstance(text, str):
        return False
    return bool(ISSUE_REF_RE.search(text) or REPO_ISSUE_REF_RE.search(text))

def has_hash_reference(text):
    if not isinstance(text, str):
        return False
    return bool(FULL_HASH_RE.search(text) or GIT_TREE_HASH_RE.search(text))

def has_log_reference(text):
    if not isinstance(text, str):
        return False
    return bool(LOG_LINE_RE.search(text))

def has_benchmark_reference(text):
    if not isinstance(text, str):
        return False
    return bool(BENCHMARK_RE.search(text))

def replace_technical_references(text):
    if not isinstance(text, str):
        return ""

    # line-wise replacement for logs / benchmark-like output
    lines = text.splitlines()
    cleaned_lines = []

    for line in lines:
        stripped = line.strip()

        if BENCHMARK_RE.match(stripped):
            cleaned_lines.append(BENCHMARK_TAG)
            continue

        if LOG_LINE_RE.match(stripped):
            cleaned_lines.append(LOG_TAG)
            continue

        cleaned_lines.append(line)

    text = "\n".join(cleaned_lines)

    # repo issue refs before plain issue refs
    text = REPO_ISSUE_REF_RE.sub(REPO_ISSUE_REF_TAG, text)
    text = ISSUE_REF_RE.sub(ISSUE_REF_TAG, text)

    # hashes / git tree hashes
    text = GIT_TREE_HASH_RE.sub(HASH_TAG, text)
    text = FULL_HASH_RE.sub(HASH_TAG, text)
    text = SHORT_HASH_RE.sub(HASH_TAG, text)

    text = re.sub(r'[ \t]+', ' ', text)
    return text.strip()

print("Technical reference rules ready")


# ---

# Crear funciones principales de normalización y limpieza
def normalize_linebreaks_and_tabs(text):
    if not isinstance(text, str):
        return ""

    # Preserve line structure for fenced blocks, diffs, quotes, and paragraph boundaries.
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = text.replace("\t", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r" *\n *", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text

def normalize_text(text):
    if not isinstance(text, str):
        return ""

    # collapse spaces and tabs, but preserve paragraph and quote separation
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = text.replace("\t", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r" *\n *", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()

print("Normalization functions ready")

def clean_code_related_text(text):
    if not isinstance(text, str) or not text.strip():
        return ""

    text = clean_inline_code(text)
    text = replace_fenced_blocks(text)
    text = replace_diff_structures(text)
    text = replace_error_logs(text)
    text = normalize_text(text)


    return text

print("Main code cleaning function ready")

def clean_comment_text(text, author=""):
    if not isinstance(text, str) or not text.strip():
        return ""

    text = COMMIT_LINE_REFERENCE_RE.sub(
        "/commit_line_reference",
        text
    )

    text = clean_bot_messages(text, author)
    text = normalize_linebreaks_and_tabs(text)
    text = clean_urls(text)
    text = replace_html_and_visual_artifacts(text)
    text = replace_ci_and_coverage_reports(text)
    text = replace_technical_references(text)
    text = clean_inline_code(text)
    text = clean_inline_code_without_backticks(text)
    text = clean_code_suggestions(text)
    text = clean_code_related_text(text)
    _, text = _extract_quotes(text)
    text = clean_approval_notifications(text)
    text = clean_cc_user_mentions(text)
    text = clean_workflow_commands(text)
    text = replace_emoticons_and_emojis(text)
    text = clean_mentions_by_paragraph(text)
    text = _anonymize_inline_mentions(text)
    text = normalize_text(text)

    return text

print("Full cleaning function ready")


# ---

# Useful discourse and social signals to preserve in text, but also track as features
WIP_SIGNAL_RE = re.compile(r'(?i)\bWIP\b')
REVIEW_SIGNAL_RE = re.compile(
    r'(?i)\b(PTAL|review required|please review|can you review|could you review|take a look|can you check|could you check|pinging)\b'
)
AGREEMENT_SIGNAL_RE = re.compile(r'(?i)(?<!\w)\+1(?!\w)|\bIIUC\b')
UNCERTAINTY_SIGNAL_RE = re.compile(
    r'(?i)\b(I am not sure|I\'m not sure|not sure|maybe|perhaps|I think|IIUC|do you mean|could it be)\b'
)
FRUSTRATION_SIGNAL_RE = re.compile(
    r'(?i)\b(frustrating|exhausting|annoying|painful|blocked|stuck|this is hard|this is difficult)\b'
)
HELP_REQUEST_SIGNAL_RE = re.compile(
    r'(?i)\b(can you help|could you help|need help|any help|could someone help|help me understand)\b'
)
URGENCY_SIGNAL_RE = re.compile(
    r'(?i)\b(asap|urgent|urgently|blocking|time-sensitive|high priority)\b'
)
POLITENESS_SIGNAL_RE = re.compile(
    r'(?i)\b(thanks|thank you|please|appreciate it|much appreciated|sorry)\b'
)
DISAGREEMENT_SIGNAL_RE = re.compile(
    r'(?i)\b(I disagree|I don\'t think|this is not correct|I don\'t agree|not convinced|I\'m not convinced|I think this is wrong)\b'
)

def has_wip_signal(text):
    return bool(WIP_SIGNAL_RE.search(text)) if isinstance(text, str) else False

def has_review_signal(text):
    return bool(REVIEW_SIGNAL_RE.search(text)) if isinstance(text, str) else False

def has_agreement_signal(text):
    return bool(AGREEMENT_SIGNAL_RE.search(text)) if isinstance(text, str) else False

def has_uncertainty_signal(text):
    return bool(UNCERTAINTY_SIGNAL_RE.search(text)) if isinstance(text, str) else False

def has_frustration_signal(text):
    return bool(FRUSTRATION_SIGNAL_RE.search(text)) if isinstance(text, str) else False

def has_help_request_signal(text):
    return bool(HELP_REQUEST_SIGNAL_RE.search(text)) if isinstance(text, str) else False

def has_urgency_signal(text):
    return bool(URGENCY_SIGNAL_RE.search(text)) if isinstance(text, str) else False

def has_politeness_signal(text):
    return bool(POLITENESS_SIGNAL_RE.search(text)) if isinstance(text, str) else False

def has_disagreement_signal(text):
    return bool(DISAGREEMENT_SIGNAL_RE.search(text)) if isinstance(text, str) else False

print("Useful discourse signal functions ready")

# ---


# URL handling: preserve anchor text and keep URL signal

URL_TAG = "/url_reference"

MARKDOWN_LINK_RE = re.compile(r'\[([^\]]+)\]\((https?://[^\s)]+|www\.[^\s)]+)\)')
GENERIC_MARKDOWN_LINK_RE = re.compile(
    r'\[([^\]]+)\]\(([^)\s]+)\)'
)
BROKEN_MARKDOWN_LINK_RE = re.compile(r'\[([^\]]+)\]\((https?:/[^\s)]+)\)')
RAW_URL_RE = re.compile(r"""(?i)\b(?:https?://|www\.)[^\s<>()\[\]{}"']+""")
BROKEN_HTTP_RE = re.compile(r"""(?i)\bhttps?:/[^\s<>()\[\]{}"']+""")

def split_trailing_punctuation(url):
    trailing = ""
    while url and url[-1] in ".,;:!?)]}":
        trailing = url[-1] + trailing
        url = url[:-1]
    return url, trailing

def _markdown_anchor_with_url(anchor):
    anchor = re.sub(r'\s+', ' ', anchor).strip()
    if not anchor:
        return URL_TAG
    return f"{anchor} {URL_TAG}"

def replace_markdown_links(text):
    if not isinstance(text, str):
        return ""

    def repl(match):
        anchor = match.group(1)
        return _markdown_anchor_with_url(anchor)

    return MARKDOWN_LINK_RE.sub(repl, text)


def replace_broken_markdown_links(text):
    if not isinstance(text, str):
        return ""

    def repl(match):
        anchor = match.group(1)
        return _markdown_anchor_with_url(anchor)

    return BROKEN_MARKDOWN_LINK_RE.sub(repl, text)


def replace_generic_markdown_links(text):
    if not isinstance(text, str):
        return ""

    def repl(match):
        anchor = match.group(1).strip()
        return _markdown_anchor_with_url(anchor)

    return GENERIC_MARKDOWN_LINK_RE.sub(repl, text)

    def repl(match):
        anchor = match.group(1)
        return _markdown_anchor_with_url(anchor)

    return BROKEN_MARKDOWN_LINK_RE.sub(repl, text)

def replace_raw_urls(text):
    if not isinstance(text, str):
        return ""

    def repl(match):
        url = match.group(0)
        core, trailing = split_trailing_punctuation(url)
        return f" {URL_TAG}{trailing} "

    text = RAW_URL_RE.sub(repl, text)
    text = BROKEN_HTTP_RE.sub(repl, text)
    text = re.sub(r'[ \t]+', ' ', text)
    return text.strip()

def has_raw_url(text):
    if not isinstance(text, str):
        return False
    return bool(RAW_URL_RE.search(text) or BROKEN_HTTP_RE.search(text))

def has_markdown_link(text):
    if not isinstance(text, str):
        return False
    return bool(MARKDOWN_LINK_RE.search(text) or BROKEN_MARKDOWN_LINK_RE.search(text))

def clean_urls(text):
    if not isinstance(text, str):
        return ""
    text = replace_markdown_links(text)
    text = replace_broken_markdown_links(text)
    text = replace_generic_markdown_links(text)
    text = replace_raw_urls(text)
    return text

print("URL handling rules ready")


# ---


# HTML, markup, and visual artifacts

HTML_DETAILS_TAG = "/html_details_block"
HTML_IMAGE_TAG = "/html_image_reference"
HTML_LINK_TAG = "/html_link_reference"
HTML_META_TAG = "/html_meta_comment"
VISUAL_ARTIFACT_TAG = "/visual_artifact_reference"

HTML_DETAILS_RE = re.compile(r'(?is)<details\b.*?>.*?</details>')
HTML_IMG_RE = re.compile(r'(?is)<img\b[^>]*>')
HTML_A_RE = re.compile(r'(?is)<a\b[^>]*href\s*=\s*["\']?[^"\'>\s]+[^>]*>(.*?)</a>')
HTML_META_COMMENT_RE = re.compile(r'(?is)<!--\s*.*?\s*-->')

VISUAL_ARTIFACT_RE = re.compile(
    r'(?im)^\s*Visible:\s*\d+%\s*-\s*\d+%\s*$'
)

CLA_CHECK_TAG = "/automated_cla_check"

CLA_CHECK_RE = re.compile(
    r'(?is)The committers listed above are authorized under a signed CLA\..*?(?=\n|$)'
)
GENERIC_HTML_TAG_RE = re.compile(
    r'(?is)</?(?:details|summary|sub|sup|br|hr|p|div|span|table|tr|td|th|thead|tbody|img|a|ul|li|ol)\b[^>]*>'
)

def has_html_markup(text):
    if not isinstance(text, str):
        return False
    return bool(
        HTML_DETAILS_RE.search(text)
        or HTML_IMG_RE.search(text)
        or HTML_A_RE.search(text)
        or HTML_META_COMMENT_RE.search(text)
        or GENERIC_HTML_TAG_RE.search(text)
    )

def has_visual_artifact(text):
    if not isinstance(text, str):
        return False
    return bool(VISUAL_ARTIFACT_RE.search(text))

def replace_html_and_visual_artifacts(text):
    if not isinstance(text, str):
        return ""

    text = HTML_DETAILS_RE.sub(f" {HTML_DETAILS_TAG} ", text)
    text = HTML_IMG_RE.sub(f" {HTML_IMAGE_TAG} ", text)
    text = CLA_CHECK_RE.sub(f" {CLA_CHECK_TAG} ", text)

    def repl_a(match):
        anchor_text = re.sub(r'\s+', ' ', match.group(1)).strip()
        if anchor_text:
            return f"{anchor_text} {HTML_LINK_TAG}"
        return HTML_LINK_TAG

    text = HTML_A_RE.sub(repl_a, text)
    text = HTML_META_COMMENT_RE.sub(f" {HTML_META_TAG} ", text)
    text = VISUAL_ARTIFACT_RE.sub(VISUAL_ARTIFACT_TAG, text)
    text = GENERIC_HTML_TAG_RE.sub(" ", text)
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r' *\n *', '\n', text)
    return text.strip()

print("HTML and visual artifact rules ready")


# ---

test = "/html_image_reference The committers listed above are authorized under a signed CLA.<ul><li>:white_check_mark: login: ermias19 / name: ermias mulugeta (/hash_reference)</li></ul>"

print(replace_html_and_visual_artifacts(test))

# ---

# CI, coverage, and automated report handling (automatic lines only)

CI_STATUS_TAG = "/ci_status_reference"
COVERAGE_REPORT_TAG = "/coverage_report_reference"

CI_STATUS_LINE_RE = re.compile(
    r'(?im)^\s*(?:CI|V8 CI|CITGM):\s*.+$'
)

CODECOV_STATUS_LINE_RE = re.compile(
    r'(?im)^\s*Codecov\b.*$'
)

COVERALLS_STATUS_LINE_RE = re.compile(
    r'(?im)^\s*Coveralls\b.*$'
)

def has_ci_status_line(text):
    if not isinstance(text, str):
        return False
    return bool(CI_STATUS_LINE_RE.search(text))

def has_codecov_report(text):
    if not isinstance(text, str):
        return False
    return bool(CODECOV_STATUS_LINE_RE.search(text))

def has_coverage_report(text):
    if not isinstance(text, str):
        return False
    return bool(COVERALLS_STATUS_LINE_RE.search(text))

def replace_ci_and_coverage_reports(text):
    if not isinstance(text, str):
        return ""

    lines = text.splitlines()
    cleaned_lines = []

    for line in lines:
        stripped = line.strip()

        if CI_STATUS_LINE_RE.match(stripped):
            cleaned_lines.append(CI_STATUS_TAG)
            continue

        # Strict automatic report lines only
        if CODECOV_STATUS_LINE_RE.match(stripped) or COVERALLS_STATUS_LINE_RE.match(stripped):
            cleaned_lines.append(COVERAGE_REPORT_TAG)
            continue

        cleaned_lines.append(line)

    return "\n".join(cleaned_lines)

print("CI, coverage, and automated report rules ready")


# ---

# Pruebas rápidas de limpieza para menciones cc y emoticones
for test in [
    "cc @klueska as well :)",
    "/cc @bart0sh",
]:
    print("ORIGINAL:", test)
    print("CLEANED :", clean_comment_text(test))
    print()


# ---

examples_mentions = [
    "@naoki9911",
    "@naoki9911 @lolita123",
    "@alice can you review this?",
    "@alice I updated the docs and release notes.",
    "I agree with @alice on this implementation.",
    "@nodejs/tsc",
    "@nodejs/tsc @nodejs/security-wg",
    "@nodejs/tsc please review",
    "@nodejs/tsc I pushed the fix.",
    "We should ask @nodejs/security-wg about this.",
    "> @dima-kov I think",
    "> @alice can you review this?",
    "> I agree with @alice on this implementation.",
    "> @nodejs/security-wg I pushed the fix."
]

for text in examples_mentions:
    print("=" * 80)
    print("ORIGINAL:", text)
    print("CLEANED :", clean_comment_text(text))


# ---

examples_emo = [
    "Looks good :)",
    "This is failing :(",
    "Thanks 👍",
    "We probably need a release note also 👍",
    "Not sure :/",
    "Nice catch 😂",
    "Great work <3"
]

for text in examples_emo:
    print("=" * 70)
    print("ORIGINAL:", text)
    print("CLEANED :", replace_emoticons_and_emojis(text))
    print("HAS    :", has_emoticon_or_emoji(text))
    print("COUNT  :", count_emoticon_or_emoji(text))
    print("TYPES  :", detect_emotion_types(text))

# ---

# Apply cleaning to the dataset with a single integrated final column

SOURCE_COL = "comment_body_raw"
AUTHOR_COL = "comment_author"
FINAL_TARGET_COL = "comment_body_clean_final"

# Auxiliary columns for rule-specific inspection
RAW_URL_TARGET_COL = "comment_body_raw_urls_only"
RAW_URL_FLAG_COL = "has_raw_url"
MARKDOWN_TARGET_COL = "comment_body_markdown_links_only"
MARKDOWN_FLAG_COL = "has_markdown_link"
COMBINED_URL_TARGET_COL = "comment_body_clean_final_urls"
CI_TARGET_COL = "comment_body_ci_reports_only"
EMOTION_TARGET_COL = "comment_body_emotions_clean"
HTML_TARGET_COL = "comment_body_html_artifacts_only"
HTML_FLAG_COL = "has_html_markup"
VISUAL_ARTIFACT_FLAG_COL = "has_visual_artifact"
TECH_REF_TARGET_COL = "comment_body_technical_refs_only"
HAS_ISSUE_REF_COL = "has_issue_ref"
HAS_HASH_REF_COL = "has_hash_reference"
HAS_LOG_REF_COL = "has_log_reference"
HAS_BENCHMARK_REF_COL = "has_benchmark_reference"
HAS_WIP_SIGNAL_COL = "has_wip_signal"
HAS_REVIEW_SIGNAL_COL = "has_review_signal"
HAS_AGREEMENT_SIGNAL_COL = "has_agreement_signal"
HAS_UNCERTAINTY_SIGNAL_COL = "has_uncertainty_signal"
HAS_FRUSTRATION_SIGNAL_COL = "has_frustration_signal"
HAS_HELP_REQUEST_SIGNAL_COL = "has_help_request_signal"
HAS_URGENCY_SIGNAL_COL = "has_urgency_signal"
HAS_POLITENESS_SIGNAL_COL = "has_politeness_signal"
HAS_DISAGREEMENT_SIGNAL_COL = "has_disagreement_signal"

# Main integrated output
df[FINAL_TARGET_COL] = df.apply(
    lambda row: clean_comment_text(
        str(row[SOURCE_COL]) if pd.notna(row[SOURCE_COL]) else "",
        str(row[AUTHOR_COL]) if AUTHOR_COL in df.columns and pd.notna(row[AUTHOR_COL]) else ""
    ),
    axis=1
)

# Rule-specific auxiliary outputs
df[RAW_URL_TARGET_COL] = df[SOURCE_COL].fillna("").astype(str).apply(replace_raw_urls)
df[RAW_URL_FLAG_COL] = df[SOURCE_COL].fillna("").astype(str).apply(has_raw_url)
df[MARKDOWN_TARGET_COL] = df[SOURCE_COL].fillna("").astype(str).apply(replace_markdown_links).apply(replace_broken_markdown_links)
df[MARKDOWN_FLAG_COL] = df[SOURCE_COL].fillna("").astype(str).apply(has_markdown_link)
df[COMBINED_URL_TARGET_COL] = df[FINAL_TARGET_COL].fillna("").astype(str).apply(clean_urls)
df[HTML_TARGET_COL] = df[SOURCE_COL].fillna("").astype(str).apply(replace_html_and_visual_artifacts)
df[TECH_REF_TARGET_COL] = df[SOURCE_COL].fillna("").astype(str).apply(replace_technical_references)
df[CI_TARGET_COL] = df[SOURCE_COL].fillna("").astype(str).apply(replace_ci_and_coverage_reports)

# Flags / features
df[HTML_FLAG_COL] = df[SOURCE_COL].fillna("").astype(str).apply(has_html_markup)
df[VISUAL_ARTIFACT_FLAG_COL] = df[SOURCE_COL].fillna("").astype(str).apply(has_visual_artifact)
df["has_ci_status_line"] = df[SOURCE_COL].fillna("").astype(str).apply(has_ci_status_line)
df["has_codecov_report"] = df[SOURCE_COL].fillna("").astype(str).apply(has_codecov_report)
df["has_coverage_report"] = df[SOURCE_COL].fillna("").astype(str).apply(has_coverage_report)

df["has_command"] = df[SOURCE_COL].fillna("").astype(str).apply(lambda x: len(extract_slash_commands(x)) > 0)
df["command_count"] = df[SOURCE_COL].fillna("").astype(str).apply(lambda x: len(extract_slash_commands(x)))
df["command_types"] = df[SOURCE_COL].fillna("").astype(str).apply(lambda x: ",".join(classify_slash_commands(x)))
df["is_slash_only"] = df[SOURCE_COL].fillna("").astype(str).apply(is_slash_only)

df["is_bot_or_automation"] = df.apply(
    lambda row: is_bot_or_automation(
        str(row[AUTHOR_COL]) if AUTHOR_COL in df.columns and pd.notna(row[AUTHOR_COL]) else "",
        str(row[SOURCE_COL]) if pd.notna(row[SOURCE_COL]) else ""
    ),
    axis=1
)

df[EMOTION_TARGET_COL] = df[SOURCE_COL].fillna("").astype(str).apply(replace_emoticons_and_emojis)
df["has_emoticon_or_emoji"] = df[SOURCE_COL].fillna("").astype(str).apply(has_emoticon_or_emoji)
df["emoticon_or_emoji_count"] = df[SOURCE_COL].fillna("").astype(str).apply(count_emoticon_or_emoji)
df["emotion_types"] = df[SOURCE_COL].fillna("").astype(str).apply(lambda x: ",".join(detect_emotion_types(x)))

df["bot_type"] = df.apply(
    lambda row: classify_bot_or_automation(
        str(row[AUTHOR_COL]) if AUTHOR_COL in df.columns and pd.notna(row[AUTHOR_COL]) else "",
        str(row[SOURCE_COL]) if pd.notna(row[SOURCE_COL]) else ""
    ),
    axis=1
)

# Main preview: original vs final integrated output
df[[SOURCE_COL, AUTHOR_COL, FINAL_TARGET_COL]].head(10)


# ---

# View changed rows using the integrated final column
SOURCE_COL = "comment_body_raw"
FINAL_TARGET_COL = "comment_body_clean_final"
CHANGED_COL = "clean_changed"

# reconstruir columnas finales
df[FINAL_TARGET_COL] = df[SOURCE_COL].fillna("").astype(str).apply(clean_comment_text)
df[CHANGED_COL] = (
    df[SOURCE_COL].fillna("").astype(str) !=
    df[FINAL_TARGET_COL].fillna("").astype(str)
)

# mover clean y changed justo al lado de raw
cols = df.columns.tolist()

for col in [FINAL_TARGET_COL, CHANGED_COL]:
    if col in cols:
        cols.remove(col)

raw_idx = cols.index(SOURCE_COL)
cols.insert(raw_idx + 1, FINAL_TARGET_COL)
cols.insert(raw_idx + 2, CHANGED_COL)

df = df[cols]

# vista rápida
df[[SOURCE_COL, FINAL_TARGET_COL, CHANGED_COL]].head(10)

# ---

# Preview the integrated final output together with key diagnostic features
preview_cols = [
    "comment_body_raw",
    "comment_author",
    "comment_body_clean_final",
    "clean_changed",
    "has_raw_url",
    "has_markdown_link",
    "has_ci_status_line",
    "has_command",
    "is_bot_or_automation",
    "has_emoticon_or_emoji",
    "has_issue_ref",
    "has_hash_reference",
    "has_log_reference",
    "has_benchmark_reference",
]

preview_cols = [c for c in preview_cols if c in df.columns]
df[preview_cols].head(20)


# ---

examples_urls = [
    "See https://example.com).",
    "Visit https://example.com, please.",
    "Check https://example.com]; it matters.",
    "Go to www.example.com?",
    "Reference: https://example.com/path/to/page).",
    "Please check [the documentation](https://docs.djangoproject.com/en/5.0/)",
    "See [ticket 123](https:/code.djangoproject.com/ticket/123)",
    "test.zip](https:/.com/user-attachments/files/21728828/test.zip)"
]

for text in examples_urls:
    print("=" * 80)
    print("ORIGINAL:", text)
    print("CLEANED :", clean_urls(text))


# ---


examples_html = [
    '<details><summary>More</summary>Hidden content</details>',
    '<img src="https://example.com/image.png" alt="preview" />',
    '<a href="https://example.com/docs">documentation</a>',
    '<!-- META -->',
    'Visible: 0% - 100%',
    '<div>Wrapped <span>content</span></div>'
]

for text in examples_html:
    print("=" * 80)
    print("ORIGINAL:", text)
    print("HAS HTML:", has_html_markup(text))
    print("HAS VISUAL ARTIFACT:", has_visual_artifact(text))
    print("CLEANED:", replace_html_and_visual_artifacts(text))


# ---

# Ver solo filas con URLs crudas detectadas

raw_url_changed = df[df["has_raw_url"]][["comment_body_raw", "comment_body_raw_urls_only", "has_raw_url"]]

print("Rows with raw URLs:", raw_url_changed.shape)
raw_url_changed.head(20)

# ---

# ======================================
# DEJAR COLUMNAS HASTA comment_body_clean_final
# ======================================

LAST_COLUMN = "comment_body_clean_final"

# posición de la última columna útil
last_idx = df.columns.get_loc(LAST_COLUMN)

# conservar columnas hasta esa posición
df_final = df.iloc[:, :last_idx + 1].copy()

print(df_final.shape)
print(df_final.columns.tolist())

# ---

# ==========================================================
# EXPORTAR DATASET FINAL CON FORMATO VISUAL
# ==========================================================

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

output_dir = Path.cwd() / "outputs"
output_dir.mkdir(parents=True, exist_ok=True)

output_path = output_dir / "dataset_clean_final.xlsx"

# Exportar dataset
df_final.to_excel(output_path, index=False)

# Abrir el Excel para aplicar formato
wb = load_workbook(output_path)
ws = wb.active

# ==========================================================
# AJUSTAR ANCHO DE COLUMNAS Y FORMATO
# ==========================================================

for col in ws.columns:
    max_len = 0
    col_letter = col[0].column_letter

    for cell in col:

        # Encabezado
        if cell.row == 1:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            cell.font = Font(bold=True)

        # Datos
        else:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

        if cell.value is not None:
            max_len = max(max_len, len(str(cell.value)))

    # Limitar ancho máximo
    ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

# ==========================================================
# ALTURA DE FILAS
# ==========================================================

# Encabezado
ws.row_dimensions[1].height = 25

# Filas de datos
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 60

# Guardar archivo
wb.save(output_path)

print("=" * 60)
print("DATASET FINAL EXPORTADO")
print("=" * 60)
print(f"Archivo guardado en: {output_path}")
print(f"Filas: {ws.max_row - 1}")
print(f"Columnas: {ws.max_column}")

# ---

pruebas = [
    "/test",
    "/test pull-kubernetes-node-kubelet-serial-crio",
    "/test pull-kubernetes-node-kubelet-serial-crio\n\n:eyes:"
]

for comentario in pruebas:
    print("=" * 70)
    print("ORIGINAL:")
    print(repr(comentario))

    print("\nSOLO WORKFLOW:")
    print(repr(clean_workflow_commands(comentario)))

    print("\nSOLO EMOJIS:")
    print(repr(replace_emoticons_and_emojis(comentario)))

    print("\nLIMPIEZA FINAL:")
    print(repr(clean_comment_text(comentario)))